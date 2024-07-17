import os
import openpyxl
import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import sys

def get_data_from_xls(file_path):
    try:
        book = xlrd.open_workbook(file_path)
        sheet = book.sheet_by_name("6.表设计")
        data = {}
        for row_idx in range(2, 7):
            col_idx = 2  # C列的索引是2
            if row_idx < sheet.nrows and col_idx < sheet.ncols:
                cell_value = sheet.cell(row_idx, col_idx).value
                if sheet.cell(row_idx, col_idx).ctype == xlrd.XL_CELL_DATE:
                    date_value = xlrd.xldate_as_tuple(cell_value, book.datemode)
                    cell_value = f"{date_value[0]}-{date_value[1]:02}-{date_value[2]:02}"
                data[f"C{row_idx + 1}"] = cell_value
            else:
                data[f"C{row_idx + 1}"] = None
        return data
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return None

def get_data_from_xlsx(file_path):
    try:
        book = openpyxl.load_workbook(file_path, data_only=True)
        sheet = book["6.表设计"]
        data = {}
        for row_idx in range(3, 8):
            col_idx = 3  # C列的索引是3
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            data[f"C{row_idx}"] = cell_value
        return data
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return None

def extract_data_from_excel(file_path):
    if file_path.endswith(".xls"):
        return get_data_from_xls(file_path)
    elif file_path.endswith(".xlsx"):
        return get_data_from_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file type: {file_path}")

def extract_name_from_filename(file_name):
    # 获取文件名不包括扩展名
    base_name = os.path.splitext(file_name)[0]
    
    # 找到最后一个 "-" 或 "_" 的索引
    last_dash_index = base_name.rfind('-')
    last_underscore_index = base_name.rfind('_')
    
    # 取最大的索引值，即最右边的分隔符
    last_separator_index = max(last_dash_index, last_underscore_index)
    
    # 提取名字部分
    if last_separator_index != -1:
        name = base_name[last_separator_index + 1:]
    else:
        name = base_name

    return name

def process_file(file_path):
    data = extract_data_from_excel(file_path)
    if data:
        file_name = os.path.basename(file_path)
        last_modified_time = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
        batch_number = file_name.split('-')[0] if '-' in file_name else ""
        name = extract_name_from_filename(file_name)
        return (
            file_name,
            data.get("C3"),
            data.get("C4"),
            data.get("C5"),
            data.get("C6"),
            data.get("C7"),
            last_modified_time,
            batch_number,
            name
        )
    return None

def progress_bar(current, total, bar_length=50):
    progress = current / total
    block = int(bar_length * progress)
    percentage = progress * 100
    text = f"\rProgress: [{'#' * block}{'-' * (bar_length - block)}] {percentage:.2f}%"
    sys.stdout.write(text)
    sys.stdout.flush()

def main(folder_path):
    results = []
    file_paths = []

    for root, _, files in os.walk(folder_path):
        for file in files:
            if (file.endswith(".xls") or file.endswith(".xlsx")) and not file.startswith("~$"):
                file_path = os.path.join(root, file)
                file_paths.append(file_path)

    total_files = len(file_paths)
    processed_files = 0

    # 使用 ThreadPoolExecutor 进行并行处理
    with ThreadPoolExecutor(max_workers=8) as executor:  # 8 可以根据您的机器配置进行调整
        future_to_file = {executor.submit(process_file, file_path): file_path for file_path in file_paths}
        for future in as_completed(future_to_file):
            result = future.result()
            if result:
                results.append(result)
            processed_files += 1
            progress_bar(processed_files, total_files)

    print()  # 打印一个新行，以结束进度条的显示

    # 提取文件夹名称以作为文件名的一部分
    folder_name = os.path.basename(os.path.normpath(folder_path))
    output_file_name = f"模型内容获取_{folder_name}.xlsx"

    # 获取脚本所在目录路径
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, output_file_name)

    # 保存到新的Excel文件中
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 添加标题行并应用样式
    headers = ["文件名", "主题对象", "一级主题", "二级主题", "英文名", "中文名", "最后修改时间", "批次号", "名字"]
    header_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 绿色底纹
    header_font = Font(color="FFFFFF", size=10, bold=True)  # 白色10号字体

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = 20  # 设置列宽为20

    for result in results:
        ws.append(result)

    wb.save(output_path)
    print(f"已获取文件数量: {len(results)}")
    print(f"文件保存路径: {output_path}")

if __name__ == "__main__":
    folder_path = "/path/to/your/folder"  # 将此路径修改为实际文件夹路径
    main(folder_path)

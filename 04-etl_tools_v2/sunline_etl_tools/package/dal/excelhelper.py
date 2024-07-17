# -*- coding:utf-8 -*-
import sys
import datetime
import os.path

import openpyxl
import xlrd
from xlutils.copy import copy
#import importlib
sys.path.append(r'D:\vscode\sunline_etl_tools\package')
#importlib.import_module(r'D:\vscode\sunline_etl_tools\package\utils')

from utils import log

print(sys.path)

_logger = log.get_logger

class ExcelRW:
    type = ""
    path = ""
    data = {}

    def __init__(self,path):
        
        self.path = path
        self.type = self.get_excel_type()
        self.data = self.get_data()

    def get_excel_type(self):
        file_extension = os.path.splitext(self.path)[1]
        if file_extension == '.xls':
            excel_type = "2003"
        elif file_extension == ".xlsx" or file_extension == ".xlsm":
            excel_type = "2007"
        else:
            excel_type = "unknown"
        return excel_type
    def get_data(self):
        if self.type == "2003":
            excel = Xls(self.path)
        elif self.type == "2007":
            excel = Xlsx(self.path)
        else:
            _logger.error("unknow excel version,exit!,file name : %s" %self.path)

        return excel.get_data() 
    def set_data(self,data,skip_line=0):
        if self.type == "2003":
            excel = Xls(self.path)
        elif self.type == "2007":
            excel = Xlsx(self.path)
        else:
            _logger.error("unknow excel version,exit!,file name : %s" %self.path)
        return excel.set_data(data,skip_line)
    
class Xls:
    def __init__(self,path):
        self.path = path
        self.book = xlrd.open_workbook(self.path)
        self.data - Xls.get_book_data(self.path)
    @staticmethod
    def get_book_data(book):
        book_dict = {}
        for sheet_name in book.sheet_names():
            sheet = book.sheet_by_name(sheet_name)
            book_dict[sheet_name] = Xls.get_sheet_data(sheet)
        return book_dict
    @staticmethod
    def get_sheet_data(sheet):
        row_count = sheet.nrows
        col_count = sheet.nrols
        data_list = []
        #循环读取行数据
        for i in range(0,row_count):
            row = []
            #读取行内各列
            for j  in range(0,col_count):
                #对数据为日期类型的进行处理
                if sheet.cell(i, j).ctype ==3:
                    date_value = xlrd.xldate_as_tuple(sheet.cell(i,j).value,0)
                    formated_date = datetime.date( *date_value[:3]).strftime('%Y-%m-%d')
                    row.append(formated_date)
                else:
                    row.append(sheet.cell(i,j).value)
            data_list.append(row)
        return data_list
    def get_data(self):
        return self.data 
    
    def set_data(self,data,skip_row):
        try:
            wb = copy(self.book)
            for sheet_name,info in data.items():
                sheet = wb.get_sheet(sheet_name)
                for i in range(len(info)):
                    row = info[i]
                    for j in range(len(row)):
                        column = row[j]
                        sheet.write(i + skip_row,j,column)

            wb.save(self.path)
            return True
        except Exception as e:
            _logger.error("Write error: %s" %e )
            return False
        
class Xlsx:
    def __init__(self,path):
        self.path = path
        self.book = openpyxl.load_workbook(self.path,data_only=True)
        self.data = Xlsx.get_book_data(self.book)
    @staticmethod
    def get_book_data(book):
        book_dict = {}
        for sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            book_dict[sheet_name] = Xlsx.get_sheet_data(sheet)
        return book_dict
    @staticmethod
    def get_sheet_data(sheet):
        row_count = sheet.max_row
        col_count = sheet.max_column
        data_list = []
        #循环读取行数据
        for i in range(1,row_count + 1):
            row = []
            #读取行内各列
            for j  in range(1,col_count + 1):
                #对数据为日期类型的进行处理
                row.append(sheet.cell(row = i,column = j).value)
            data_list.append(row)
        return data_list
    def get_data(self):
        return self.data 
    
    def set_data(self,data,skip_line):
        return True

if __name__ == "__main__":
    file_path = r"D:\vscode\sunline_etl_tools\icl_c_pt_corp_cust_perf_rela.xlsx"      
    excel = ExcelRW(file_path)
    excel.get_data()
    print(excel.data.keys())
    print(excel.data["icl_c_pt_corp_cust_perf_rela"][10])
    print(excel.data["icl_c_pt_corp_cust_perf_rela"][10][4])     


#!/usr/bin/env python3
# -*- coding:utf-8 -*-

"""
作者: zjj
日期: 20231128
功能: 用于生成基于CEA模板的sql文件，sql文件会生成至 ${ELT_HOME}/autocode/dml/icl/ 目录下
使用方法：python %ETL_HOME%/script/gen_cea_sql.py 归属层次 模板文件路径
        例：python %ETL_HOME%/script/gen_cea_sql.py icl D:/tmp/test.xlsx
日志:
    20231128 zjj v1.0
"""
import os
import sys

from jinja2 import Environment, FileSystemLoader
import openpyxl


sys.path.append('D:\\sunline_etl_tools')

from package.utils import excelhelper, filehelper
from package.cea import cea_sdm_reader

etl_home = os.environ["ETL_HOME"]


def render(file, info):
    filepath, filename = os.path.split(file)
    env = Environment(loader=FileSystemLoader(filepath))
    template = env.get_template(filename)
    content = template.render(info)
    return content.replace("\n    \n", "\n").replace("\n        \n", "\n")


def main():
    # 获取命令行参数
    args = sys.argv
    if len(args) < 3:
        print("Usage: python gen_cea_sql.py <归属层次> <模板文件路径>")
        return
    at_level, file_path = args[1], args[2]
    print(at_level)
    print(file_path)
    print(etl_home)

    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(file_path)

    # 获取 index sheet 页
    index_sheet = workbook['index']

    # 遍历 A 列和 B 列的数据，从第三行开始
    for row in index_sheet.iter_rows(min_row=3, values_only=True):
        data_a = row[2]
        data_b = row[12]
        data_c = row[13]

        # 判断 B 列的值是否为 Y
        if data_b == 'Y':
            # 根据 A 列的数据获取对应的 sheet 页
            if data_a in workbook.sheetnames:
                sheet_name = data_a
                template_file = "%s/templage/cea/cea_template.sql" % etl_home
                sdm_excel = excelhelper.Xlsx(file_path)
                sdm_dict = sdm_excel.data
                sdm_martix = sdm_dict[sheet_name]
                cea_sdm = cea_sdm_reader.CeaSDM(sdm_martix)
                if cea_sdm:
                    #print(render(template_file, cea_sdm.sdm))
                    output_cea_sdm_file = "%s/autocode/sum/%s.sql" % (etl_home, sheet_name.lower())
                    contnent_cea_sdm = render(template_file, cea_sdm.sdm)
                    filehelper.write_file(output_cea_sdm_file, contnent_cea_sdm)
                    
                else:
                    print("Unable to render SQL file for sheet: %s" % sheet_name)

    # 关闭 Excel 文件
    workbook.close()


if __name__ == '__main__':
    main()

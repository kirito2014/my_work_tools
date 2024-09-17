import os
import threading
import tkinter as tk
from tkinter import filedialog, ttk
from ttkthemes import ThemedTk
from pathlib import Path
import pandas as pd
import shutil
from openpyxl import load_workbook
from datetime import datetime

#pyinstaller --onefile --noconsole --add-data "res;res" --icon=sunline.ico case_split.py  # 打包命令
# 模块1: 读取文件并提取部门信息,用于获取部门列表 方便循环操作
def get_unique_departments(source_file: str, sheet_name: str, usecols:str) -> list:
    """从源文件中读取部门信息并去重"""
    df = pd.read_excel(source_file, sheet_name=sheet_name)
    departments = df[usecols].dropna().unique().tolist()  # 去重部门列表 
    departments = [str(dp).strip() for dp in departments if dp != '合计']  # 去除空格
    #print(departments)
    return departments

# 模块2: 检测并删除旧文件
def check_and_remove_file(new_file_name: str,output_directory: str):
    """检查文件夹下是否有该文件，如有则删除"""
    # 拼接文件名
    new_file_name = os.path.join(output_directory, new_file_name)
    #检查文件夹是否存在该文件，如有则删除
    if os.path.exists(new_file_name):
        os.remove(new_file_name)
        #print(f"文件 {os.path.basename(new_file_name)} 已删除。")


# 模块3: 创建新文件并保存到output文件夹
def create_new_file(template_file: str, new_file_name: str,output_directory: str):
    """从模板创建新的部门文件"""
    new_file_name = os.path.join(output_directory, new_file_name)
    shutil.copy(template_file, new_file_name)  # 复制模板文件
    #print(f"新文件 {os.path.basename(new_file_name)} 已创建。")
    app.info_label.config(text=f"[ INFO ] 新文件 {os.path.basename(new_file_name)} 已创建.", foreground="#298073")


# 模块4: 读取部门数据并筛选写入新文件
def filter_department_data(workbook_name: str, sheet_name: str, department_name: str,usecols:str):
    # 检查文件是否存在
    if not os.path.exists(workbook_name):
        raise FileNotFoundError(f"文件 '{workbook_name}' 不存在")
    # 检查工作簿是否能被打开，以及是否包含指定的 sheet
    try:
        excel_file = pd.ExcelFile(workbook_name)
    except Exception as e:
        raise ValueError(f"无法读取文件 '{workbook_name}'，错误: {e}")
    if sheet_name not in excel_file.sheet_names:
        raise ValueError(f"Sheet '{sheet_name}' 在文件 '{workbook_name}' 中不存在")
    # 读取 Excel 文件中的指定 sheet
    df = pd.read_excel(workbook_name, sheet_name=sheet_name)
    
    # 检查 筛选 列是否存在 todo 修改筛选判断逻辑
    if usecols not in df.columns:
        raise ValueError(f"{usecols} 列在工作表中不存在")
    # 筛选 B 列中对应部门名称的数据
    filtered_df = df[df[usecols] == department_name]
    return filtered_df

# 模块5: 获取当前月份并返回yyyymm字符串
def get_current_month() -> str:
    """获取当前月份并返回yyyymm字符串"""
    from datetime import datetime
    current_date = datetime.now()
    current_year = current_date.year
    current_month = current_date.month
    file_date = f"{current_year}{current_month:02d}"
    return file_date

# 模块6: sheet页与列名对应关系
cols_mapping = {
    '售前情况统计表': '业务部',
    '部门商机明细': '归属业务部',
    '投标数据': '归属业务部',
    '商机报工明细':'归属业务部门'
}
# 模块7:自动获取对应列名
def get_usecols(sheet_name: str):
    column_name = cols_mapping.get(sheet_name)
    
    # 如果没有找到匹配的列名，返回 None
    if column_name is None:
        app.info_label.config(text=f"[WARNING] 无法找到 sheet '{sheet_name}' 对应的列名.", foreground="#DB231D")
    
    return column_name
# 模块8: 拼接新文件名，模板名_部门名_当前月份，模板名以_分割 去除第二个元素，拼接新文件名
def get_new_file_name(template_file: str, dp_name: str):
    """拼接新文件名，模板名_部门名_当前月份，模板名以_分割 去除第二个元素，拼接新文件名"""
    file_date = get_current_month()
    file_name = os.path.basename(template_file)
    file_name_parts = file_name.split('_')
    new_file_name = f"{file_name_parts[0]}_{dp_name}_{file_date}.xlsx"
    return new_file_name
# 模块9: 写入特定单元格

def write_to_specific_cells(file_path: str, sheet_name: str,dp_name: str):
    # 打开 Excel 文件
    wb = load_workbook(file_path)
    
    if sheet_name not in wb.sheetnames:
        app.info_label.config(text=f"[WARNING] 工作表 '{sheet_name}' 不存在于文件中.", foreground="#DB231D")
        return
    ws = wb[sheet_name]
    # 写入 C1 单元格
    ws['C1'] = dp_name
    # 写入 D4 单元格
    ws['D4'] = dp_name
    # 写入 D5 单元格，当前时间
    ws['D5'] = datetime.now().strftime('%Y-%m-%d')
    #写入 D6 单元格，(yyyy年mm月）
    ws['D6'] = datetime.now().strftime('%Y年%m月')

    # 保存更改
    wb.save(file_path)
    app.info_label.config(text=f"[ INFO ] 文件 {file_path} 已更新.", foreground="#298073")




class App():
    def __init__(self, root):
        self.root = root
        self.root.title("解决方案部-售前统计工具")
        self.root.geometry('500x200')
        self.root.configure(bg='#f0f0f0')
        self.root.set_theme("arc")
        self.root.option_add("*Font", "黑体 10")
        self.load_logo()

        self.source_file_var = tk.StringVar()
        self.template_file_var = tk.StringVar()

        frame = tk.Frame(self.root, bg='#f0f0f0')
        frame.pack(pady=30, padx=50, anchor="e")
        
        ttk.Button(frame, text="选择源文件", command=self.select_folder, width=20).grid(row=0, column=0, padx=10, pady=5)
        ttk.Button(frame, text="选择模板文件", command=self.select_target_file, width=20).grid(row=0, column=1, padx=10, pady=5)
        ttk.Button(frame, text="确认执行", command=self.run_script, width=20).grid(row=1, column=0, padx=10, pady=5)
        ttk.Button(frame, text="清除信息", command=self.clear_info, width=20).grid(row=1, column=1, padx=10, pady=5)

        # 创建进度条
        self.progress = ttk.Progressbar(self.root, orient='horizontal', mode='determinate', length=320)
        self.progress.pack(padx=50,pady=10)

        self.info_label = ttk.Label(self.root, text="请选择文件", foreground="#DB231D", background="#f0f0f0")
        self.info_label.pack()

    def load_logo(self):
        # 省略 logo 加载部分的代码
        pass

    def select_folder(self):
        source_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        ##print(source_path)
        self.clear_info()
        if source_path:
            self.source_file_var.set(source_path)
            self.update_info_label(source_path)
        else:
            self.info_label.config(text="请选择要处理的源文件", foreground="#F24405")

    def select_target_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if file_path:
            self.template_file_var.set(file_path)
            self.update_info_label(file_path)
        else:
            self.info_label.config(text="请选择模板文件", foreground="#F24405")

    def update_info_label(self, path):
        self.info_label.config(text=f"已选择：{os.path.basename(path)}", foreground='#298073')

    def run_script(self):
            self.progress['value'] = 0  # 重置进度条
            source_file_path = self.source_file_var.get()
            target_file_path = self.template_file_var.get()

            if not source_file_path or not Path(source_file_path).is_file():
                self.info_label.config(text=f"[ERROR] 请选择有效源文件.", foreground="#DB231D")
                return

            if not target_file_path or not Path(target_file_path).is_file():
                self.info_label.config(text=f"[ERROR] 请选择有效模板文件.", foreground="#DB231D")
                return

            threading.Thread(target=self.process_files, args=(source_file_path, target_file_path, "售前情况输出")).start()

    def process_files(self, source_file_path, target_file_path, output_directory):
        try:
            # 检查输出文件夹是否存在，没有则新建文件夹
            output_path = os.path.join(os.path.dirname(source_file_path), output_directory)
            if not os.path.exists(output_path):
                os.makedirs(output_path)
                #print(f"文件夹 {output_path} 已创建。")
                self.info_label.config(text=f"[INFO] 文件夹 {output_path} 已创建.", foreground="#298073")
                

            # 获取所有不重复的部门名称
            departments = get_unique_departments(source_file_path, sheet_name='售前情况统计表', usecols=get_usecols('售前情况统计表'))
            
            total_steps = len(departments)
            step_size = 100 / total_steps  # 计算每个部门的进度步进

            for index, dp_name in enumerate(departments):
                new_file_name = get_new_file_name(target_file_path, dp_name)
                #print(new_file_name)
                # 检查并删除旧文件
                check_and_remove_file(new_file_name, output_path)
                create_new_file(target_file_path, new_file_name, output_path)
                # 循环 cols_mapping 的值作为 sheet_name
                self.info_label.config(text=f"[INFO] 正在处理 <{dp_name}> .", foreground="#298073")
                for sheet_name in cols_mapping.keys():
                    target_file = os.path.join(output_path, new_file_name)
                    filter_col = get_usecols(sheet_name)
                    data = filter_department_data(source_file_path, sheet_name=sheet_name, department_name=dp_name, usecols=filter_col)
                    # 将筛选后的数据写入目标文件的指定 sheet_name 的 A1 单元格
                    with pd.ExcelWriter(target_file, mode='a', if_sheet_exists='overlay') as writer:
                        data.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                    # 填写目标文件 标题及目录 sheet 页
                write_to_specific_cells(target_file, sheet_name="标题及目录", dp_name=dp_name)

                self.update_progress(step_size * (index + 1))  # 更新进度条

            self.info_label.config(text=f"[INFO] 文件合并处理完成.", foreground="#298073")
        except Exception as e:
            self.info_label.config(text=f"[ERROR] 执行脚本失败: {e}", foreground="#DB231D")


    def update_progress(self, value):
        """更新进度条"""
        self.progress['value'] = value
        self.root.update_idletasks()

    def clear_info(self):
        self.source_file_var.set("")
        self.template_file_var.set("")
        self.progress['value'] = 0  # 重置进度条

if __name__ == "__main__":
    try:
        root = ThemedTk(theme=False)
        app = App(root)
        root.mainloop()
    except Exception as e:
        app.info_label.config(text=f"[ERROR] 出现错误: {e}", foreground="#DB231D")
        

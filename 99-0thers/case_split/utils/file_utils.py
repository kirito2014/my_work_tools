import os
import pandas as pd
import shutil
from openpyxl import load_workbook
from datetime import datetime


# 模块1: 读取文件并提取部门信息,用于获取部门列表 方便循环操作
def get_unique_departments(source_file: str, sheet_name: str, usecols:str) -> list:
    """从源文件中读取部门信息并去重"""
    df = pd.read_excel(source_file, sheet_name=sheet_name)
    departments = df[usecols].dropna().unique().tolist()  # 去重部门列表 
    departments = [str(dp).strip() for dp in departments if dp != '合计']  # 去除空格
    print(departments)
    return departments

# 模块2: 检测并删除旧文件
def check_and_remove_file(new_file_name: str,output_directory: str):
    """检查文件夹下是否有该文件，如有则删除"""
    # 拼接文件名
    new_file_name = os.path.join(output_directory, new_file_name)
    #检查文件夹是否存在该文件，如有则删除
    if os.path.exists(new_file_name):
        os.remove(new_file_name)
        print(f"文件 {os.path.basename(new_file_name)} 已删除。")


# 模块3: 创建新文件并保存到output文件夹
def create_new_file(template_file: str, new_file_name: str,output_directory: str):
    """从模板创建新的部门文件"""
    new_file_name = os.path.join(output_directory, new_file_name)
    shutil.copy(template_file, new_file_name)  # 复制模板文件
    print(f"新文件 {os.path.basename(new_file_name)} 已创建。")


# 模块4: 读取部门数据并筛选写入新文件
def filter_department_data(workbook_name: str, sheet_name: str, department_name: str,usecols:str):
    # 检查文件是否存在
    if not os.path.exists(workbook_name):
        raise FileNotFoundError(f"文件 '{workbook_name}' 不存在")
    # 检查工作簿是否能被打开，以及是否包含指定的 sheet
    try:
        excel_file = pd.ExcelFile(workbook_name)
    except Exception as e:
        raise ValueError(f"无法读取文件 '{workbook_name}'，错误: {e}")
    if sheet_name not in excel_file.sheet_names:
        raise ValueError(f"Sheet '{sheet_name}' 在文件 '{workbook_name}' 中不存在")
    # 读取 Excel 文件中的指定 sheet
    df = pd.read_excel(workbook_name, sheet_name=sheet_name)
    
    # 检查 筛选 列是否存在 todo 修改筛选判断逻辑
    if usecols not in df.columns:
        raise ValueError(f"{usecols} 列在工作表中不存在")
    # 筛选 B 列中对应部门名称的数据
    filtered_df = df[df[usecols] == department_name]
    return filtered_df

# 模块5: 获取当前月份并返回yyyymm字符串
def get_current_month() -> str:
    """获取当前月份并返回yyyymm字符串"""
    from datetime import datetime
    current_date = datetime.now()
    current_year = current_date.year
    current_month = current_date.month
    file_date = f"{current_year}{current_month:02d}"
    return file_date

# 模块6: sheet页与列名对应关系
cols_mapping = {
    '售前情况统计表': '业务部',
    '部门商机明细': '归属业务部',
    '投标数据': '归属业务部',
    '商机报工明细':'归属业务部门'
}
# 模块7:自动获取对应列名
def get_usecols(sheet_name: str):
    column_name = cols_mapping.get(sheet_name)
    
    # 如果没有找到匹配的列名，返回 None
    if column_name is None:
        print(f"Warning: 无法找到 sheet '{sheet_name}' 对应的列名")
    
    return column_name
# 模块8: 拼接新文件名，模板名_部门名_当前月份，模板名以_分割 去除第二个元素，拼接新文件名
def get_new_file_name(template_file: str, dp_name: str):
    """拼接新文件名，模板名_部门名_当前月份，模板名以_分割 去除第二个元素，拼接新文件名"""
    file_date = get_current_month()
    file_name = os.path.basename(template_file)
    file_name_parts = file_name.split('_')
    new_file_name = f"{file_name_parts[0]}_{dp_name}_{file_date}.xlsx"
    return new_file_name
# 模块9: 写入特定单元格

def write_to_specific_cells(file_path: str, sheet_name: str,dp_name: str):
    # 打开 Excel 文件
    wb = load_workbook(file_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' 不存在于文件中")
        return
    ws = wb[sheet_name]
    # 写入 C1 单元格
    ws['C1'] = dp_name
    # 写入 D4 单元格
    ws['D4'] = dp_name
    # 写入 D5 单元格，当前时间
    ws['D5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # 保存更改
    wb.save(file_path)
    print(f"文件 {file_path} 已更新")

# 主函数: 完整流程
def process_files(template_file: str, source_file: str, output_directory: str):
    """循环处理部门列表，创建新文件"""

    #检查输出文件夹是否存在，没有则新建文件夹
    #输出文件夹为输入文件夹所在目录 拼接output_directory
    output_path = os.path.join(os.path.dirname(source_file), output_directory)
    if not os.path.exists(output_path):
        os.makedirs(output_path)
        print(f"文件夹 {output_path} 已创建。")

    #获取所有不重复的部门名称
    departments = get_unique_departments(source_file, sheet_name='售前情况统计表',usecols=get_usecols('售前情况统计表'))

    for dp_name in departments:
 
        new_file_name = get_new_file_name(template_file,dp_name)
        print(new_file_name)
        #检查并删除旧文件
        check_and_remove_file(new_file_name,output_path)
        create_new_file(template_file, new_file_name,output_path)
        #循环cols_mapping 的值作为sheet_name
        for sheet_name in cols_mapping.keys():
            #打开目标文件
            target_file = os.path.join(output_path, new_file_name)
            filter_col = get_usecols(sheet_name)
            data = filter_department_data(source_file, sheet_name=sheet_name, department_name=dp_name, usecols=filter_col)
            # 将筛选后的数据写入目标文件的指定 sheet_name 的 A1 单元格
            with pd.ExcelWriter(target_file, mode='a', if_sheet_exists='overlay') as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
        # 填写目标文件 标题及目录sheet页  C1 单元格为sheet_name ,D4单元格为sheet_name D5单元格为当前时间
        write_to_specific_cells(target_file, sheet_name="标题及目录", dp_name=dp_name)



if __name__ == '__main__':

    template_file = 'H:/python/tools/my_work_tools/99-0thers/case_split/case_file/售前统计月报_模板.xlsx'
    source_file = 'H:/python/tools/my_work_tools/99-0thers/case_split/case_file/售前统计_202408(1).xlsm'
    output_directory = 'output_case_split'
    try:
        #result_df = filter_department_data(workbook_name, sheet_name, department_name)
        #print(result_df)
        process_files(template_file, source_file, output_directory)
    except Exception as e:
        print(f"发生错误: {e}")





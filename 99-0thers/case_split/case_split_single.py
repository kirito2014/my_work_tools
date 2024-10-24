# -*- coding: utf-8 -*-
"""
********************************************************************
*
*    Filename   :  case_split_single.py
*    Description:  split excel file by department
*
*    Version    :  1.1.2
*    Created    :  2024/09/28 10:25:07
*    updated   :  2024/10/20 10:25:07
*    Compiler   :  python
*
*    Author     :  wangmujun(解决方案部/战略规划部), 
*    Company    :  深圳市长亮科技股份有限公司
*    E-mail     :  wangmujun@sunline.cn
*    License    :  GPL
*
*    Copyright (c) 2024, Sunline Corporation. All rights reserved.
*
********************************************************************
"""
import os,sys
import threading
import tkinter as tk
#import pywinstyles
from tkinter import filedialog, ttk
from ttkthemes import ThemedTk
from pathlib import Path
import pandas as pd
import shutil
from openpyxl import load_workbook
from datetime import datetime
from PIL import Image, ImageTk
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment

#pyinstaller --onefile --noconsole --add-data "res;res" --icon=sunline.ico case_split.py  # 打包命令
# 模块1: 读取文件并提取部门信息,用于获取部门列表 方便循环操作
def get_unique_departments(source_file: str, sheet_name: str, usecols:str) -> list:
    """从源文件中读取部门信息并去重"""
    df = pd.read_excel(source_file, sheet_name=sheet_name)
    departments = df[usecols].dropna().unique().tolist()  # 去重部门列表 
    departments = [str(dp).strip() for dp in departments if dp != '合计']  # 去除空格
    #print(departments)
    return departments

# 模块2: 检测并删除旧文件
def check_and_remove_file(new_file_name: str,output_directory: str):
    """检查文件夹下是否有该文件，如有则删除"""
    # 拼接文件名
    new_file_name = os.path.join(output_directory, new_file_name)
    #检查文件夹是否存在该文件，如有则删除
    if os.path.exists(new_file_name):
        os.remove(new_file_name)
        #print(f"文件 {os.path.basename(new_file_name)} 已删除。")


# 模块3: 创建新文件并保存到output文件夹
def create_new_file(template_file: str, new_file_name: str,output_directory: str):
    """从模板创建新的部门文件"""
    new_file_name = os.path.join(output_directory, new_file_name)
    shutil.copy(template_file, new_file_name)  # 复制模板文件
    #print(f"新文件 {os.path.basename(new_file_name)} 已创建。")
    app.info_label.config(text=f"[ INFO ] 新文件 {os.path.basename(new_file_name)} 已创建.", foreground="#298073")


# 模块4: 读取部门数据并筛选写入新文件
def filter_department_data(workbook_name: str, sheet_name: str, department_name: str,usecols:str):
    # 检查文件是否存在
    if not os.path.exists(workbook_name):
        raise FileNotFoundError(f"文件 '{workbook_name}' 不存在")
    # 检查工作簿是否能被打开，以及是否包含指定的 sheet
    try:
        excel_file = pd.ExcelFile(workbook_name)
    except Exception as e:
        raise ValueError(f"无法读取文件 '{workbook_name}'，错误: {e}")
    if sheet_name not in excel_file.sheet_names:
        raise ValueError(f"Sheet '{sheet_name}' 在文件 '{workbook_name}' 中不存在")
    # 读取 Excel 文件中的指定 sheet
    df = pd.read_excel(workbook_name, sheet_name=sheet_name)
    
    # 检查 筛选 列是否存在 todo 修改筛选判断逻辑
    if usecols not in df.columns:
        raise ValueError(f"{usecols} 列在工作表中不存在")
    # 筛选 B 列中对应部门名称的数据
    filtered_df = df[df[usecols] == department_name]
    return filtered_df

# 模块5: 获取当前月份并返回yyyymm字符串
def get_current_month() -> str:
    """获取当前月份并返回yyyymm字符串"""
    from datetime import datetime
    current_date = datetime.now()
    current_year = current_date.year
    current_month = current_date.month
    file_date = f"{current_year}{current_month:02d}"
    return file_date

# 模块6: sheet页与列名对应关系
cols_mapping = {
    '售前情况统计表': '业务部',
    '部门商机明细': '归属业务部',
    '投标数据': '归属业务部',
    '商机报工明细':'归属业务部门'
}
# 模块7:自动获取对应列名
def get_usecols(sheet_name: str):
    column_name = cols_mapping.get(sheet_name)
    
    # 如果没有找到匹配的列名，返回 None
    if column_name is None:
        app.info_label.config(text=f"[WARNING] 无法找到 sheet '{sheet_name}' 对应的列名.", foreground="#DB231D")
    
    return column_name
# 模块8: 拼接新文件名，模板名_部门名_当前月份，模板名以_分割 去除第二个元素，拼接新文件名
def get_new_file_name(template_file: str, dp_name: str):
    """拼接新文件名，模板名_部门名_当前月份，模板名以_分割 去除第二个元素，拼接新文件名"""
    file_date = get_current_month()
    file_name = os.path.basename(template_file)
    file_name_parts = file_name.split('_')
    new_file_name = f"{file_name_parts[0]}_{dp_name}_{file_date}.xlsx"
    return new_file_name
# 模块9: 写入特定单元格

def write_to_specific_cells(file_path: str, sheet_name: str,dp_name: str):
    # 打开 Excel 文件
    wb = load_workbook(file_path)
    #如果是目录sheet页填入内容，如果不是则执行美化
    if sheet_name == "标题及目录":    
        if sheet_name not in wb.sheetnames:
            app.info_label.config(text=f"[WARNING] 工作表 '{sheet_name}' 不存在于文件中.", foreground="#DB231D")
            return
        ws = wb[sheet_name]
        # 写入 C1 单元格
        ws['C1'] = dp_name
        # 写入 D4 单元格
        ws['D4'] = dp_name
        # 写入 D5 单元格，当前时间
        ws['D5'] = datetime.now().strftime('%Y-%m-%d')
        #写入 D6 单元格，(yyyy年mm月）
        ws['D6'] = datetime.now().strftime('%Y年%m月')
    else:
        try:
            sheet = wb[sheet_name]

            # 设置字体、边框和对齐方式
            font = Font(name='宋体', size=10)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            alignment = Alignment(vertical='center')

            # 获取当前 sheet 页的最大行和最大列
            max_row = sheet.max_row
            max_col = sheet.max_column

            # 从 A2 开始遍历整个数据区域，添加边框，修改字体，设置垂直居中
            for row in sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.font = font
                    cell.border = border
                    cell.alignment = alignment

            # 保存文件
            wb.save(file_path)
            app.info_label.config(text=f"[INFO] Sheet {sheet_name} 已美化.", foreground="#298073")

        except Exception as e:
            app.info_label.config(text=f"[ERROR] 美化 sheet {sheet_name} 时发生错误: {e}")
        finally:
            wb.save(file_path)
            wb.close()
    # 保存更改
    wb.save(file_path)
    app.info_label.config(text=f"[ INFO ] 文件 {os.path.basename(file_path)} 已更新首页并调格式.", foreground="#298073")



#模块10:logo展示用获取打包后资源文件的路径
def get_resource_path(relative_path):
    """获取打包后资源文件的路径"""
    try:
        # 获取应用程序目录
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

#模块11:文件内容美化,写入标题内容
def beautify_sheet(self, file_path, sheet_name):
    """对指定的 sheet 页进行美化，添加边框、设置字体和垂直居中"""
    try:
        # 使用 openpyxl 打开文件
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        # 设置字体、边框和对齐方式
        font = Font(name='宋体', size=10)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment = Alignment(vertical='center')

        # 获取当前 sheet 页的最大行和最大列
        max_row = sheet.max_row
        max_col = sheet.max_column

        # 从 A2 开始遍历整个数据区域，添加边框，修改字体，设置垂直居中
        for row in sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.font = font
                cell.border = border
                cell.alignment = alignment

        # 保存文件
        wb.save(file_path)
        self.info_label.config(text=f"[INFO] Sheet {sheet_name} 已美化.", foreground="#298073")

    except Exception as e:
        self.info_label.config(text=f"[ERROR] 美化 sheet {sheet_name} 时发生错误: {e}")
    finally:
        wb.save(file_path)
        wb.close()

#主窗口
class App():
    def __init__(self, root):
        self.root = root
        self.root.title("解决方案部-售前统计工具")
        self.root.geometry('600x800')
        self.root.configure(bg='#f0f0f0')
        self.root.set_theme("arc")
        self.load_logo()

        self.source_file_var = tk.StringVar()
        self.template_file_var = tk.StringVar()
        self.department_vars = []
        self.department_checkbuttons = []
        self.select_all_var = tk.BooleanVar()  # 全选控制变量

        # 调用布局方法
        self.create_layout()
    def create_layout(self):
        # 左侧区域 - 部门复选框区域，带有滚动条
        left_frame = tk.Frame(self.root, bg='#f0f0f0', width=150, height=300)
        left_frame.pack(side='left', fill='both', padx=10, pady=30, expand=True)

        self.scroll_canvas = tk.Canvas(left_frame, bg='#f0f0f0', width=150, height=300)
        self.scrollbar = tk.Scrollbar(left_frame, orient="vertical", command=self.scroll_canvas.yview)
        self.scroll_canvas.configure(yscrollcommand=self.scrollbar.set)
        self.scrollbar.pack(side='left', fill='y')
        self.scroll_canvas.pack(side='left', fill='both', expand=True)
        #滚动条宽度25，颜色为蓝色,圆角样式
        self.scrollbar.config(width=25, troughcolor="#f0f0f0", activebackground="#298073", highlightthickness=0, bd=0)

        self.scrollable_frame = tk.Frame(self.scroll_canvas, bg='#f0f0f0')
        self.scroll_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.scrollable_frame.bind("<Configure>", lambda e: self.scroll_canvas.configure(scrollregion=self.scroll_canvas.bbox("all")))

        # 右侧区域 - 按钮区域,
        right_frame = tk.Frame(self.root, bg='#f0f0f0', width=150, height=300)
        right_frame.pack(side='top', fill='y', padx=10, pady=80)

        #竖向分割线，样式为虚线，颜色为灰色
        ttk.Separator(left_frame, orient='vertical').pack(side='right', fill='y', padx=20, pady=10)

        # 按钮设置
        ttk.Button(right_frame, text="选择来源文件", command=self.select_folder, width=20).pack(pady=5)
        ttk.Button(right_frame, text="选择模板文件", command=self.select_target_file, width=20).pack(pady=5)
        #ttk.Button(right_frame, text="确认全量执行", command=self.run_script, width=20).pack(pady=5)
        ttk.Button(right_frame, text="获取部门列表", command=self.load_departments, width=20).pack(pady=5)
        ttk.Button(right_frame, text="执行文件拆分", command=self.run_selected_departments, width=20).pack(pady=5)
        ttk.Button(right_frame, text="清除执行信息", command=self.clear_info, width=20).pack(pady=5)

        self.select_all_checkbox = None  # 延迟显示的全选按钮

        # 底部区域 - 进度条和信息展示
        bottom_frame = tk.Frame(self.root, bg='#f0f0f0', width=200, height=80)
        bottom_frame.pack(side='bottom', fill='x', padx=10, pady=10)
        #横向分割线，样式为虚线，颜色为灰色
        ttk.Separator(bottom_frame, orient='horizontal').pack(side='top', fill='x', padx=10, pady=5)

        self.progress = ttk.Progressbar(bottom_frame, orient='horizontal', mode='determinate', length=290)
        self.progress.pack(side='bottom',fill='x', padx=10, pady=20)

        self.info_label = ttk.Label(bottom_frame, text="请选择文件", foreground="#DB231D", background="#f0f0f0")
        self.info_label.pack()

        self.info_label_select_source = ttk.Label(right_frame, text=f"使用方法：\n\t<--选择要拆分的文件 \n\t<--选择拆分模板文件 \n\t<--获取部门列表\n\t<--选择拆分部门或全选\n\t<--执行拆分\r\n程序版权所有：深圳市长亮科技有限公司", foreground="#298073", background="#f0f0f0")
        self.info_label_select_source.pack(pady=5)

    def toggle_select_all(self):
        """控制全选/取消全选功能"""
        select_all = self.select_all_var.get()
        for var in self.department_vars:
            var.set(select_all)

    def load_logo(self):
        logo_path = get_resource_path('res/sunline_logo_original.png')  # 请确保路径正确
        logo_image = Image.open(logo_path)
        logo_image = logo_image.resize((int(logo_image.width * 0.4), int(logo_image.height * 0.4)), Image.LANCZOS)
        self.logo_photo = ImageTk.PhotoImage(logo_image)

        # 显示logo在右上角
        self.logo_label = ttk.Label(self.root, image=self.logo_photo, background='#f0f0f0')
        self.logo_label.place(x=410, y=20)  # 设置图片位置

    def select_folder(self):
        source_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        ##print(source_path)
        self.clear_info()
        if source_path:
            self.source_file_var.set(source_path)
            self.update_info_label(source_path)
        else:
            self.info_label.config(text="请选择要处理的源文件", foreground="#F24405")

    def select_target_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if file_path:
            self.template_file_var.set(file_path)
            self.update_info_label(file_path)
        else:
            self.info_label.config(text="请选择模板文件", foreground="#F24405")

    def update_info_label(self, path):
        self.info_label.config(text=f"已选择：{os.path.basename(path)}", foreground='#298073')

    def run_script(self):
            self.progress['value'] = 0  # 重置进度条
            source_file_path = self.source_file_var.get()
            target_file_path = self.template_file_var.get()

            if not source_file_path or not Path(source_file_path).is_file():
                self.info_label.config(text=f"[ERROR] 请选择有效源文件.", foreground="#DB231D")
                return

            if not target_file_path or not Path(target_file_path).is_file():
                self.info_label.config(text=f"[ERROR] 请选择有效模板文件.", foreground="#DB231D")
                return

            threading.Thread(target=self.process_files, args=(source_file_path, target_file_path, "售前情况输出")).start()

    def load_departments(self):
        """从源文件中获取部门列表并生成复选框"""
        source_file_path = self.source_file_var.get()
        if not source_file_path or not Path(source_file_path).is_file():
            self.info_label.config(text="请先选择有效的源文件", foreground="#DB231D")
            return

        departments = get_unique_departments(source_file_path, sheet_name='售前情况统计表', usecols=get_usecols('售前情况统计表'))

        # 清除旧的复选框
        for cb in self.department_checkbuttons:
            cb.pack_forget()
        self.department_vars = []
        self.department_checkbuttons = []

        # 生成复选框
        for department in departments:
            var = tk.BooleanVar()
            self.department_vars.append(var)
            cb = ttk.Checkbutton(self.scrollable_frame, text=department, variable=var)
            cb.pack(anchor='w')  # 放置到窗口上
            self.department_checkbuttons.append(cb)

        # 显示“全选”按钮
        if not self.select_all_checkbox:
            self.select_all_checkbox = ttk.Checkbutton(self.scrollable_frame, text="全选", variable=self.select_all_var, command=self.toggle_select_all)
            self.select_all_checkbox.pack(anchor='w', before=self.department_checkbuttons[0])  # 显示在列表最上方

        self.info_label.config(text=f"[INFO] 获取部门列表成功，选择要执行的部门并执行.", foreground="#298073")


    def process_selected_departments( self,source_file_path, target_file_path, output_directory, selected_departments):
        """处理选中的部门"""
        try:
            total_steps = len(selected_departments)
            step_size = 100 / total_steps  # 计算每个部门的进度步进
            for index,department in enumerate(selected_departments):
            
                self.process_single_department(source_file_path, target_file_path, output_directory, department)
                self.update_progress(step_size * (index + 1))  # 更新进度条
            self.info_label.config(text=f"[ INFO ] 拆分文件处理完成", foreground="#298073")
        except Exception as e:
            self.info_label.config(text=f"[ERROR] 处理失败: {e}", foreground="#DB231D")

    def run_selected_departments(self):
        """执行选中的部门的文件生成"""
        source_file_path = self.source_file_var.get()
        target_file_path = self.template_file_var.get()

        if not source_file_path or not Path(source_file_path).is_file():
            self.info_label.config(text="请选择有效的源文件", foreground="#DB231D")
            return

        if not target_file_path or not Path(target_file_path).is_file():
            self.info_label.config(text="请选择有效的模板文件", foreground="#DB231D")
            return

        if self.select_all_var.get():
            # 全选状态，执行全量处理
            self.run_script()
        else:
            # 仅处理选中的部门
            selected_departments = [dep for var, dep in zip(self.department_vars, get_unique_departments(source_file_path, sheet_name='售前情况统计表', usecols=get_usecols('售前情况统计表'))) if var.get()]
            if not selected_departments:
                self.info_label.config(text="请选择至少一个部门", foreground="#DB231D")
                return
            threading.Thread(target=self.process_selected_departments, args=(source_file_path, target_file_path, "售前情况输出", selected_departments)).start()

    def process_single_department(self,source_file_path, target_file_path, output_directory, department_name):
        """处理单个部门的数据"""
        try:
            # 创建输出目录
            output_path = os.path.join(os.path.dirname(source_file_path), output_directory)
            if not os.path.exists(output_path):
                os.makedirs(output_path)
                self.info_label.config(text=f"[INFO] 文件夹 {output_path} 已创建.", foreground="#298073")

            # 生成新文件名
            new_file_name = get_new_file_name(target_file_path, department_name)
            self.info_label.config(text=f"[INFO] 处理部门: {department_name}, 文件名: {new_file_name}")

            # 删除旧文件并创建新文件
            check_and_remove_file(new_file_name, output_path)
            create_new_file(target_file_path, new_file_name, output_path)

            # 循环处理 cols_mapping 中的 sheet_name
            for sheet_name in cols_mapping.keys():
                filter_col = get_usecols(sheet_name)
                data = filter_department_data(source_file_path, sheet_name=sheet_name, department_name=department_name, usecols=filter_col)
                
                # 写入到目标文件
                target_file = os.path.join(output_path, new_file_name)
                with pd.ExcelWriter(target_file, mode='a', if_sheet_exists='overlay') as writer:
                    data.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)

                # 美化除了 "标题及目录" 之外的 sheet 页
                # if sheet_name != "标题及目录":
                #     self.beautify_sheet(target_file, sheet_name)
                # 在标题及目录 sheet 中填写部门信息
                write_to_specific_cells(target_file, sheet_name=sheet_name, dp_name=department_name)
            self.info_label.config(text=f"[INFO] 文件拆分处理完成.", foreground="#298073")
        except Exception as e:
            self.info_label.config(text=f"[ERROR] 处理部门 {department_name} 时发生错误: {e}")

    def process_files(self, source_file_path, target_file_path, output_directory):
        try:
            # 检查输出文件夹是否存在，没有则新建文件夹
            output_path = os.path.join(os.path.dirname(source_file_path), output_directory)
            if not os.path.exists(output_path):
                os.makedirs(output_path)
                #print(f"文件夹 {output_path} 已创建。")
                self.info_label.config(text=f"[INFO] 文件夹 {output_path} 已创建.", foreground="#298073")
                

            # 获取所有不重复的部门名称
            departments = get_unique_departments(source_file_path, sheet_name='售前情况统计表', usecols=get_usecols('售前情况统计表'))
            
            total_steps = len(departments)
            step_size = 100 / total_steps  # 计算每个部门的进度步进

            for index, dp_name in enumerate(departments):
                new_file_name = get_new_file_name(target_file_path, dp_name)
                #print(new_file_name)
                # 检查并删除旧文件
                check_and_remove_file(new_file_name, output_path)
                create_new_file(target_file_path, new_file_name, output_path)
                # 循环 cols_mapping 的值作为 sheet_name
                self.info_label.config(text=f"[INFO] 正在处理 <{dp_name}> .", foreground="#298073")
                for sheet_name in cols_mapping.keys():
                    target_file = os.path.join(output_path, new_file_name)
                    filter_col = get_usecols(sheet_name)
                    data = filter_department_data(source_file_path, sheet_name=sheet_name, department_name=dp_name, usecols=filter_col)
                    # 将筛选后的数据写入目标文件的指定 sheet_name 的 A1 单元格
                    with pd.ExcelWriter(target_file, mode='a', if_sheet_exists='overlay') as writer:
                        data.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                    # 填写目标文件 标题及目录 sheet 页

                    # 美化除了 "标题及目录" 之外的 sheet 页
                    # if sheet_name != "标题及目录":
                    #     self.beautify_sheet(target_file, sheet_name)
                    write_to_specific_cells(target_file, sheet_name=sheet_name, dp_name=dp_name)

                self.update_progress(step_size * (index + 1))  # 更新进度条

            self.info_label.config(text=f"[INFO] 文件拆分处理完成.", foreground="#298073")
        except Exception as e:
            self.info_label.config(text=f"[ERROR] 执行脚本失败: {e}", foreground="#DB231D") 

    def update_progress(self, value):
        """更新进度条"""
        self.progress['value'] = value
        self.root.update_idletasks()

    def clear_info(self):
        self.source_file_var.set("")
        self.template_file_var.set("")
        #清除复选框
        # 清除旧的复选框
        for cb in self.department_checkbuttons:
            cb.pack_forget()
        self.department_vars = []
        self.department_checkbuttons = []
        #清除旧的全选框
        self.select_all_var.set(False)
        if  self.select_all_checkbox:
            self.select_all_checkbox.pack_forget()

        self.progress['value'] = 0  # 重置进度条
if __name__ == "__main__":
    try:
        root = ThemedTk(theme=False)
        app = App(root)
        root.mainloop()
    except Exception as e:
        print(f"出现错误: {e}")
        input("按任意键退出...")


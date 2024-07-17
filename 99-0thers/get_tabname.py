import os
import re
import openpyxl

# 函数用于提取 SQL 文件中的表名
def extract_table_names(sql_content):
    # 使用正则表达式匹配 FROM 后的表名
    pattern = r'FROM\s+(\w+\.\w+)\s+'
    matches = re.findall(pattern, sql_content)
    return matches

# 函数用于过滤表名列表中包含文件名的表名
def filter_table_names(file_name, table_names):
    filtered_table_names = []
    for table_name in table_names:
        if table_name not in file_name:
            filtered_table_names.append(table_name)
    return filtered_table_names

# 遍历文件夹下的每个 SQL 文件
def process_folder(folder_path):
    data = {}  # 存储文件名和表名列表的字典
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".sql"):
            file_path = os.path.join(folder_path, file_name)
            with open(file_path, 'r') as file:
                sql_content = file.read()
                table_names = extract_table_names(sql_content)
                filtered_table_names = filter_table_names(file_name, table_names)
                data[file_name] = filtered_table_names
    return data

# 写入数据到 Excel 文件
def write_to_excel(data, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = '文件名'
    sheet['B1'] = '表名'
    row = 2
    for file_name, table_names in data.items():
        sheet.cell(row=row, column=1, value=file_name)
        for table_name in table_names:
            sheet.cell(row=row, column=2, value=table_name)
            row += 1
    workbook.save(output_file)

# 主函数
def main(folder_path, output_file):
    data = process_folder(folder_path)
    write_to_excel(data, output_file)

# 示例用法
if __name__ == "__main__":
    folder_path = "your_folder_path"  # 替换为实际的文件夹路径
    output_file = "output.xlsx"  # 输出 Excel 文件名
    main(folder_path, output_file)

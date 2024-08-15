# -*- coding: utf-8 -*-
"""
********************************************************************
*
*    Filename   :  case_collector.py
*
*    Description:  Demonstrate how to add copyright comments in Python.
*
*    Version    :  1.0
*    Created    :  2024/08/01 10:25:07
*    Revision   :  none
*    Compiler   :  python
*
*    Author     :  wangmujun(解决方案部/战略规划部), 
*    Company    :  深圳市长亮科技股份有限公司
*
*    Author(s)  :  wangmujun (wangmujun@sunline.cn)
*
*    License    :  GPL
*
*    Copyright (c) 2024, Sunline Corporation. All rights reserved.
*
********************************************************************
"""


import os,sys,re
import glob
import pandas as pd
import openpyxl
from collections import OrderedDict
import tkinter as tk
from tkinter import Tk, Label
from tkinter import filedialog, messagebox,Text,ttk,scrolledtext
from pathlib import Path
from PIL import Image, ImageTk
import queue
import threading

sys.stdout.reconfigure(encoding='utf-8')

def rename_file(file_path):
    # 循环文件夹下的所有文件
    for root, dirs, filenames in os.walk(file_path):
        for file_name in filenames:
            file_path = os.path.join(root, file_name)
            if not os.path.basename(file_name).startswith('~$'):#排除临时文件
                new_file_name = file_path.replace('\xa0', '').replace(' ','')
                os.rename(file_path, new_file_name)

def merge_dicts(dict_list):
    merged_dict = {}
    for d in dict_list:
        merged_dict.update(d)
    return [merged_dict]


#格式化基本信息
def format_basic_info_data(case_data):
    '''
    根据特定的单元格内容分割数据，并格式化为不同部分的 DataFrame。
    '''
    # 标志开始和结束的关键字
    start_key = "*合同名称"
    end_key = "*二、项目实施内容"
    
    start_index = None
    end_index = None
    basic_info_dict = []
    # 查找开始和结束的行索引
    for i, row in case_data.iterrows():
        #print(row.iloc[0])
        cell_value = str(row.iloc[0]).strip()  # 将单元格值转换为字符串
        #print(cell_value)
        if cell_value.startswith(start_key):
            start_index = i
        elif cell_value.startswith(end_key):
            end_index = i
            break
    #print(start_index )
    #print(end_index)
    if start_index is not None and end_index is not None:
        # 提取基本信息部分的数据
        basic_info_df = case_data.iloc[start_index:end_index].reset_index(drop=True)
        if basic_info_df is not None:
            #print("基本信息:")
            fddd = basic_info_df.iloc[1:].dropna(how='all').copy()
            fddd.columns = basic_info_df.iloc[0]
            #print(fddd)
            # 将 DataFrame 转换为数组
            data_array = fddd.to_numpy()
            column_names = fddd.columns.tolist()
            
            # 转换成字典列表
            for row in data_array:
                row_dict = {}
                for key, value in zip(column_names, row):
                    #如果value 是nan则value为/
                    if pd.isna(value):
                        value = "/"
                    if isinstance(key, str) and key.startswith("*合同名称"):
                        row_dict["contract_name"] = value
                    elif isinstance(key, str) and key.startswith("*项目规模"):
                        row_dict["project_scale"] = value
                    elif isinstance(key, str) and key.startswith("*合同年份"):
                        row_dict["contract_year"] = value
                    elif isinstance(key, str) and key.startswith("*是否人力外包"):
                        row_dict["is_outsourced"] = value
                    elif isinstance(key, str) and key.startswith("*客户简称"):
                        row_dict["cust_abb"] = value
                    elif isinstance(key, str) and key.startswith("*客户联系人及联系方式"):
                        row_dict["cust_contact"] = value
                    elif isinstance(key, str) and key.startswith("是否涉及新开业或合并"):
                        row_dict["is_new_cust"] = value
                    elif isinstance(key, str) and key.startswith("*主管系统部"):
                        row_dict["hd_managedp"] = value
                    elif isinstance(key, str) and key.startswith("*分管系统部"):
                        row_dict["asi_managedp"] = value
                    elif isinstance(key, str) and key.startswith("*项目经理"):
                        row_dict["project_manager"] = value
                    elif isinstance(key, str) and key.startswith("*销售区域"):
                        row_dict["sale_area"] = value
                    elif isinstance(key, str) and key.startswith("*实施周期"):
                        row_dict["project_period"] = value
                    elif isinstance(key, str) and key.startswith("*项目状态"):
                        row_dict["project_status"] = value
                basic_info_dict.append(row_dict)
                #print(basic_info_dict)
                return basic_info_dict
    else:
        log_message(f"[ ERROR ]无法找到 <{start_key}> 或 <{end_key}>","red")
        return None

#格式化项目实施内容,这部分用不到
def format_project_exc_data(case_data):
    '''
    根据特定的单元格内容分割数据，并格式化为不同部分的 DataFrame。
    '''
    # 标志开始和结束的关键字
    start_key = "*二、项目实施内容"
    end_key = "*三、案例分类"
    
    start_index = None
    end_index = None
    project_info_dict = []
    # 查找开始和结束的行索引
    for i, row in case_data.iterrows():
        #print(row.iloc[0])
        cell_value = str(row.iloc[0]).strip()  # 将单元格值转换为字符串
        #print(cell_value)
        if cell_value.startswith(start_key):
            start_index = i
        elif cell_value.startswith(end_key):
            end_index = i
            break
    #print(start_index )
    #print(end_index)
    if start_index is not None and end_index is not None:
        # 提取基本信息部分的数据
        project_info_df = case_data.iloc[start_index:end_index].reset_index(drop=True)
        if project_info_df is not None:
            #print("项目实施内容:")
            fddd = project_info_df.iloc[1:].dropna(how='all').copy()
            fddd.columns = project_info_df.iloc[0]
            if len(fddd) <= 1:
                default_row = ['/' for _ in range(len(fddd.columns))]
                fddd.loc[len(fddd)] = default_row
            #print(fddd)
            # 将 DataFrame 转换为数组
            data_array = fddd.to_numpy()
            column_names = fddd.columns.tolist()
            for row in data_array:
                row_dict = {}
                for key, value in zip(column_names, row):
                    if pd.isna(value):
                        value = "/"
                    if isinstance(key, str) and key.startswith("*二、项目实施内容"):
                        row_dict["project_info"] = value
                project_info_dict.append(row_dict)
                #print(basic_info_dict)
                return project_info_dict
    else:
        log_message(f"[ ERROR ]无法找到 <{start_key}> 或 <{end_key}>","red")
        return None


def format_case_classify_data(case_data):
    '''
    根据特定的单元格内容分割数据，并格式化为不同部分的 DataFrame。
    '''
    # 标志开始和结束的关键字
    start_key = "*三、案例分类"
    end_key = "*四、案例通用补充信息"
    
    start_index = None
    end_index = None
    class_info_dict = []
    # 查找开始和结束的行索引
    for i, row in case_data.iterrows():
        #print(row.iloc[0])
        cell_value = str(row.iloc[0]).strip()  # 将单元格值转换为字符串
        #print(cell_value)
        if cell_value.startswith(start_key):
            start_index = i
        elif cell_value.startswith(end_key):
            end_index = i
            break
    if start_index is not None and end_index is not None:
        # 提取基本信息部分的数据
        case_class_df = case_data.iloc[start_index:end_index].reset_index(drop=True)
        if case_class_df is not None:
            #print("案例分类信息:")
            fddd = case_class_df.iloc[2:].dropna(how='all').copy()
            fddd.columns = case_class_df.iloc[0]

            # 将 DataFrame 转换为数组
            # 根据索引值提取不同的列
            for index, row in fddd.iterrows():
                row_dict = {}
                if index > 13:
                    case_category = row.iloc[0]
                    case_name = row.iloc[7]
                else:
                    case_category = row.iloc[1]
                    case_name = row.iloc[7]
                if pd.isna(case_name):
                    case_name = "/"
                if isinstance(case_category, str) and case_category.startswith("数据架构咨询规划"):
                    row_dict["consulting_plan"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("数据平台类"):
                    row_dict["data_platform"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("配套产品实施"):
                    row_dict["product_implement"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("平台实施"):
                    row_dict["platform_implement"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("咨询"):
                    row_dict["busi_consult"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("条线集市"):
                    row_dict["line_market"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("指标/标签"):
                    row_dict["indicator"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("数据服务"):
                    row_dict["data_service"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("经营分析类"):
                    row_dict["operate_analysis"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("客户管理"):
                    row_dict["customer_management"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("风险管理"):
                    row_dict["risk_management"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("监管报送"):
                    row_dict["supervisory_report"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("业务系统"):
                    row_dict["business_platform"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("代销"):
                    row_dict["agency_sales"] = case_name
                elif isinstance(case_category, str) and case_category.startswith("其他"):
                    row_dict["others"] = case_name               
                class_info_dict.append(row_dict)
            #print(class_info_dict)
            return class_info_dict
    
    else:
        log_message(f"[ ERROR ]无法找到 <{start_key}> 或 <{end_key}>","red")
        return None

#格式化通用补充信息信息
def format_supply_info_data(case_data):
    '''
    根据特定的单元格内容分割数据，并格式化为不同部分的 DataFrame。
    '''
    # 标志开始和结束的关键字
    start_key = "*四、案例通用补充信息"
    end_key = "*五、其它补充信息"
    
    start_index = None
    end_index = None
    supply_info_dict = []
    # 查找开始和结束的行索引
    for i, row in case_data.iterrows():
        #print(row.iloc[0])
        cell_value = str(row.iloc[0]).strip()  # 将单元格值转换为字符串
        #print(cell_value)
        if cell_value.startswith(start_key):
            start_index = i
        elif cell_value.startswith(end_key):
            end_index = i
            break
    # print(start_index )
    # print(end_index)
    if start_index is not None and end_index is not None:
        # 提取数据
        supply_info_df = case_data.iloc[start_index:end_index].reset_index(drop=True)
        if supply_info_df is not None:
            #print("通用补充信息:")
    
            fddd = supply_info_df.iloc[3:].dropna(how='all').copy()
            if len(fddd) <= 1:
                default_row = ['/' for _ in range(len(fddd.columns))]
                fddd.loc[len(fddd)] = default_row

            fddd.columns = supply_info_df.iloc[2]
            fddd.reset_index(drop=True, inplace=True)
            # 将 DataFrame 转换为数组
            #print(fddd)
            data_array = fddd.to_numpy()
            #print(data_array)
            column_names = fddd.columns.tolist()
            
            # 转换成字典列表
            for row in data_array:
                row_dict = {}
                for key, value in zip(column_names, row):
                    if pd.isna(value):
                        value = "/"
                    if isinstance(key, str) and key.startswith("数据库类型"):
                        row_dict["database_type"] = value
                    elif isinstance(key, str) and key.startswith("数据库产品版本号"):
                        row_dict["database_version"] = value
                    elif isinstance(key, str) and key.startswith("节点数"):
                        row_dict["node_number"] = value
                    elif isinstance(key, str) and key.startswith("代理厂商"):
                        row_dict["agent_corpration"] = value
                    elif isinstance(key, str) and key.startswith("BI工具"):
                        row_dict["bi_tool"] = value
                    elif isinstance(key, str) and key.startswith("调度平台产品"):
                        row_dict["scheduling_platform"] = value
                    elif isinstance(key, str) and key.startswith("开发平台产品"):
                        row_dict["data_development"] = value
                    elif isinstance(key, str) and key.startswith("交换平台产品"):
                        row_dict["data_exchange"] = value
                    elif isinstance(key, str) and key.startswith("数据资产管理产品"):
                        row_dict["dataassest_product"] = value
                    elif isinstance(key, str) and key.startswith("模型管理产品"):
                        row_dict["model_product"] = value
                    elif isinstance(key, str) and key.startswith("服务器产品型号"):
                        row_dict["server_model"] = value
                    elif isinstance(key, str) and key.startswith("操作系统版本号"):
                        row_dict["operation_version"] = value
                    elif isinstance(key, str) and key.startswith("中间件版本号"):
                        row_dict["middleware_version"] = value
                supply_info_dict.append(row_dict)
                #print(supply_info_dict)
                return supply_info_dict
    else:
        log_message(f"[ ERROR ]无法找到 <{start_key}> 或 <{end_key}>","red")
        return None


#格式化其他补充信息信息
def format_other_info_data(case_data):
    '''
    根据特定的单元格内容分割数据，并格式化为不同部分的 DataFrame。
    '''
    # 标志开始和结束的关键字
    start_key = "*五、其它补充信息"
    #获取df的最大行


    start_index = None
    end_index = len(case_data)  # 结束行为数据的最大行
    other_info_dict = []
    # 查找开始和结束的行索引
    for i, row in case_data.iterrows():
        #print(row.iloc[0])
        cell_value = str(row.iloc[0]).strip()  # 将单元格值转换为字符串
        #print(cell_value)
        if cell_value.startswith(start_key):
            start_index = i
            break
    # print(start_index )
    # print(end_index)
    if start_index is not None and end_index is not None:
        # 提取数据
        other_info_df = case_data.iloc[start_index:end_index].reset_index(drop=True)
        if other_info_df is not None:
            #print("其他补充信息:")
            fddd = other_info_df.iloc[3:].dropna(how='all').copy()
            fddd.columns = other_info_df.iloc[2]
            #将第12列的列名重命名为1111
            # 如果提取的有效数据行数小于 2，添加一行默认值
            if len(fddd) <= 1:
                default_row = ['/' for _ in range(len(fddd.columns))]
                fddd.loc[len(fddd)] = default_row
            fddd.rename(columns={fddd.columns[11]: '原实施厂商'}, inplace=True)
            fddd.reset_index(drop=True, inplace=True)
            # 将 DataFrame 转换为数组
            #print(fddd)
            data_array = fddd.to_numpy()
            #print(data_array)
            column_names = fddd.columns.tolist()
            
            # 转换成字典列表
            for row in data_array:
                row_dict = {}
                for key, value in zip(column_names, row):
                    if pd.isna(value):
                        value = "/"
                    if isinstance(key, str) and key.startswith("支撑应用"):
                        row_dict["supply_application"] = value
                    elif isinstance(key, str) and key.startswith("是否配合新核心改造"):
                        row_dict["core_transformation"] = value
                    elif isinstance(key, str) and key.startswith("数据源"):
                        row_dict["data_source"] = value
                    elif isinstance(key, str) and key.startswith("是否数据迁移"):
                        row_dict["migration_flag"] = value
                    elif isinstance(key, str) and key.startswith("迁移目标数据产品"):
                        row_dict["mig_distnation"] = value
                    elif isinstance(key, str) and key.startswith("计划迁移周期"):
                        row_dict["migration_period"] = value
                    elif isinstance(key, str) and key.startswith("迁移进展"):
                        row_dict["migration_progress"] = value
                    elif isinstance(key, str) and key == "报表工具":
                        row_dict["bi_tool"] = value
                    elif isinstance(key, str) and key.startswith("是否代理报表工具"):
                        row_dict["agent_bi_tool"] = value
                    elif isinstance(key, str) and key.startswith("报表工具节点数"):
                        row_dict["bi_tool_nodes"] = value
                    elif isinstance(key, str) and key.startswith("报表工具知识库"):
                        row_dict["bi_tool_database"] = value
                    elif isinstance(key, str) and key.startswith("原实施厂商"):
                        row_dict["original_implementation"] = value
                    elif isinstance(key, str) and key.startswith("是否信创"):
                        row_dict["innovation_flag"] = value
                other_info_dict.append(row_dict)
               # print(other_info_dict)
                return other_info_dict
    else:
        log_message(f"[ ERROR ] 无法找到 <{start_key}> ","red")
        return None

# 文档检查
def check_and_process_file(filepath):
    # 打开Excel文件
    try:
        df_source = pd.read_excel(filepath, sheet_name='案例调研')
        # 格式化数据，提取各个部分的信息
        basic_info_dict = format_basic_info_data(df_source)
        project_info = format_project_exc_data(df_source)
        class_data = format_case_classify_data(df_source)
        supply_data = format_supply_info_data(df_source)
        other_data = format_other_info_data(df_source)
        if class_data:
            class_data = merge_dicts(class_data)
        # 创建一个包含第一个元素的集合
        data_first_elements = {
            "basic_info": basic_info_dict[0] if basic_info_dict else None,
            "project_info": project_info[0] if project_info else None,
            "class_data": class_data[0] if class_data else None,
            "supply_data": supply_data[0] if supply_data else None,
            "other_data": other_data[0] if other_data else None
        }
        
        return data_first_elements
    
    except Exception as e:
        log_message(f"[ ERROR ] 打开文件<{filepath}>失败: {e}","red")
        return None
    
# 获取最大非空行
def get_max_non_empty_row(sheet):
    max_row = sheet.max_row
    for row in range(max_row, 0, -1):
        for cell in sheet[row]:
            if cell.value is not None:
                return row
    return 0

#将处理好的数据加载到目标表中
def load_data_to_target_file(data, target_file_path,source_file_path):
    try:
        workbook = openpyxl.load_workbook(target_file_path)
    except Exception as e:
        log_message(f"[ ERROR ] 无法加载目标文件: {e}","red")
        return

    sheets_to_check = ["案例清单", "通用补充信息", "数据应用类补充信息","客户清单"]
    cust_list = ''

    #判断basic数据是否存在，判断客户名称是否为/

    #提前准备合同名称是所有sheet页的“主键”
    #数据加载
    basic_info = data.get("basic_info", None)
    project_info = data.get("project_info", None)
    class_data = data.get("class_data", None)
    supply_data = data.get("supply_data", None)
    other_data = data.get("other_data", None)
    if not basic_info:
        log_message(f"[ ERROR ] 未找到<{os.path.basename(source_file_path)}>中填写的基本信息",'red')
        return
    #获取基本信息中的合同名称和客户名称

    contract_name  = basic_info.get("contract_name", None)    
    cust_abb        = basic_info.get("cust_abb", None)

    #如果客户名称是/代表为空，不处理数据
    if contract_name and contract_name != "/":
    #添加到客户清单列表中
        cust_list = cust_abb

        for sheet_name in sheets_to_check:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                max_non_empty_row = get_max_non_empty_row(sheet)
                #print(max_non_empty_row)
                
                # 确保最大非空行数 =3
                if max_non_empty_row >= 3:
                    target_row = max_non_empty_row + 1

                    if sheet_name == "案例清单":
                        #基本信息
                        project_scale  = basic_info.get("project_scale", None)
                        contract_year   = basic_info.get("contract_year", None)
                        is_outsourced   = basic_info.get("is_outsourced", None)
                        cust_contact    = basic_info.get("cust_contact", None)
                        is_new_cust     = basic_info.get("is_new_cust", None)
                        hd_managedp     = basic_info.get("hd_managedp", None)
                        asi_managedp    = basic_info.get("asi_managedp", None)
                        project_manager = basic_info.get("project_manager", None)
                        sale_area       = basic_info.get("sale_area", None)
                        project_period  = basic_info.get("project_period", None)
                        project_status  = basic_info.get("project_status", None)

                        #项目实施内容
                        project_info = project_info.get("project_info", None)

                        #分类信息
                        consulting_plan 	= class_data.get("consulting_plan", None)
                        data_platform       = class_data.get("data_platform", None)
                        product_implement   = class_data.get("product_implement", None)
                        platform_implement  = class_data.get("platform_implement", None)
                        busi_consult        = class_data.get("busi_consult", None)
                        line_market         = class_data.get("line_market", None)
                        indicator           = class_data.get("indicator", None)
                        data_service        = class_data.get("data_service", None)
                        operate_analysis    = class_data.get("operate_analysis", None)
                        customer_management = class_data.get("customer_management", None)
                        risk_management     = class_data.get("risk_management", None)
                        supervisory_report  = class_data.get("supervisory_report", None)
                        business_platform   = class_data.get("business_platform", None)
                        agency_sales        = class_data.get("agency_sales", None)
                        others              = class_data.get("others", None)

                        #写入数据
                        sheet.cell(row=target_row, column=1, value=contract_name)
                        sheet.cell(row=target_row, column=2, value=project_scale)
                        sheet.cell(row=target_row, column=3, value=contract_year)
                        sheet.cell(row=target_row, column=4, value=is_outsourced)
                        sheet.cell(row=target_row, column=5, value=cust_abb)
                        sheet.cell(row=target_row, column=35, value=cust_contact)
                        sheet.cell(row=target_row, column=37, value=is_new_cust)
                        sheet.cell(row=target_row, column=10, value=hd_managedp)
                        sheet.cell(row=target_row, column=11, value=asi_managedp)
                        sheet.cell(row=target_row, column=12, value=project_manager)
                        sheet.cell(row=target_row, column=34, value=sale_area)
                        sheet.cell(row=target_row, column=29, value=project_period)
                        sheet.cell(row=target_row, column=30, value=project_status)
                        sheet.cell(row=target_row, column=28, value=project_info)

                        sheet.cell(row=target_row, column=14, value=consulting_plan)
                        sheet.cell(row=target_row, column=15, value=data_platform)
                        sheet.cell(row=target_row, column=16, value=product_implement)
                        sheet.cell(row=target_row, column=17, value=platform_implement)
                        sheet.cell(row=target_row, column=18, value=busi_consult)
                        sheet.cell(row=target_row, column=19, value=line_market)
                        sheet.cell(row=target_row, column=20, value=indicator)
                        sheet.cell(row=target_row, column=21, value=data_service)
                        sheet.cell(row=target_row, column=22, value=operate_analysis)
                        sheet.cell(row=target_row, column=23, value=customer_management)
                        sheet.cell(row=target_row, column=24, value=risk_management)
                        sheet.cell(row=target_row, column=25, value=supervisory_report)
                        sheet.cell(row=target_row, column=26, value=business_platform)
                        sheet.cell(row=target_row, column=27, value=agency_sales)
                        sheet.cell(row=target_row, column=36, value=others)
                        #填充公式

                        sheet.cell(row=target_row, column=6, value=f'=IFERROR(VLOOKUP($E{target_row},客户清单!B:G,2,0),"")')
                        sheet.cell(row=target_row, column=7, value=f'=IFERROR(VLOOKUP($E{target_row},客户清单!B:G,3,0),"")')
                        sheet.cell(row=target_row, column=8, value=f'=IFERROR(VLOOKUP($E{target_row},客户清单!B:G,4,0),"")')
                        sheet.cell(row=target_row, column=9, value=f'=IF(VLOOKUP($E{target_row},客户清单!B:G,6,0)=0,"<暂未更新>",VLOOKUP($E{target_row},客户清单!B:G,6,0))')
                        sheet.cell(row=target_row, column=31, value=f'=IFERROR(HYPERLINK("#\'通用补充信息\'!A"&MATCH(A{target_row}, 通用补充信息!A:A, 0), ">>>通用补充信息<<<"), ">>>暂无补充<<<")')
                        sheet.cell(row=target_row, column=32, value=f'=IFERROR(HYPERLINK("#\'数据应用类补充信息\'!A"&MATCH(A{target_row}, 数据应用类补充信息!A:A, 0), ">>>数据应用类补充信息<<<"), ">>>暂无补充<<<")')

                        log_message(f"[SUCCESS] 写入 {os.path.basename(source_file_path)} 的数据到 <{sheet_name}> ","green")
                        #print(f"写入 {os.path.basename(source_file_path)} 的数据到 {sheet_name} ")
                    elif sheet_name == "通用补充信息":

                        supply_application        = other_data.get("supply_application", None) 
                        core_transformation       = other_data.get("core_transformation", None) 
                        migration_flag            = other_data.get("migration_flag", None) 
                        mig_distnation            = other_data.get("mig_distnation", None) 
                        migration_period          = other_data.get("migration_period", None) 
                        migration_progress        = other_data.get("migration_progress", None) 
                        innovation_flag           = other_data.get("innovation_flag", None) 
                        database_type	      = supply_data.get("database_type", None) 
                        database_version      = supply_data.get("database_version", None) 
                        node_number           = supply_data.get("node_number", None) 
                        agent_corpration      = supply_data.get("agent_corpration", None) 
                        bi_tool               = supply_data.get("bi_tool", None) 
                        scheduling_platform   = supply_data.get("scheduling_platform", None) 
                        data_development      = supply_data.get("data_development", None) 
                        data_exchange         = supply_data.get("data_exchange", None) 
                        dataassest_product    = supply_data.get("dataassest_product", None) 
                        model_product         = supply_data.get("model_product", None) 
                        server_model          = supply_data.get("server_model", None) 
                        operation_version     = supply_data.get("operation_version", None) 
                        middleware_version    = supply_data.get("middleware_version", None) 




                        sheet.cell(row=target_row, column=1, value=contract_name)
                        sheet.cell(row=target_row, column=16, value=supply_application)
                        sheet.cell(row=target_row, column=17, value=core_transformation)
                        sheet.cell(row=target_row, column=18, value=migration_flag)
                        sheet.cell(row=target_row, column=19, value=mig_distnation)
                        sheet.cell(row=target_row, column=20, value=migration_period)
                        sheet.cell(row=target_row, column=21, value=migration_progress)
                        sheet.cell(row=target_row, column=22, value=innovation_flag)

                        sheet.cell(row=target_row, column=2, value=database_type)
                        sheet.cell(row=target_row, column=3, value=database_version)
                        sheet.cell(row=target_row, column=4, value=node_number)
                        sheet.cell(row=target_row, column=5, value=agent_corpration)
                        sheet.cell(row=target_row, column=6, value=bi_tool)
                        sheet.cell(row=target_row, column=7, value=scheduling_platform)
                        sheet.cell(row=target_row, column=8, value=data_development)
                        sheet.cell(row=target_row, column=9, value=data_exchange)
                        sheet.cell(row=target_row, column=10, value=dataassest_product)
                        sheet.cell(row=target_row, column=11, value=model_product)
                        sheet.cell(row=target_row, column=12, value=server_model)
                        sheet.cell(row=target_row, column=13, value=operation_version)
                        sheet.cell(row=target_row, column=14, value=middleware_version)
                        sheet.cell(row=target_row, column=14, value='/')

                        log_message(f"[SUCCESS] 写入 {os.path.basename(source_file_path)} 的数据到 <{sheet_name}> ","green")
                    elif sheet_name == "数据应用类补充信息":

                        data_source               = other_data.get("data_source", None)  #
                        bi_tool                   = other_data.get("bi_tool", None) #
                        agent_bi_tool             = other_data.get("agent_bi_tool", None)  #
                        bi_tool_nodes             = other_data.get("bi_tool_nodes", None) #
                        bi_tool_database          = other_data.get("bi_tool_database", None)  #
                        original_implementation   = other_data.get("original_implementation", None) #

                        sheet.cell(row=target_row, column=1, value=contract_name)
                        sheet.cell(row=target_row, column=2, value=bi_tool)#
                        sheet.cell(row=target_row, column=3, value=agent_bi_tool)#
                        sheet.cell(row=target_row, column=4, value=bi_tool_nodes)#
                        sheet.cell(row=target_row, column=5, value=bi_tool_database)#
                        sheet.cell(row=target_row, column=6, value=data_source) #
                        sheet.cell(row=target_row, column=7, value=original_implementation)#

                        log_message(f"[SUCCESS] 写入 {os.path.basename(source_file_path)} 的数据到 <{sheet_name}> ","green")

            else:
                log_message(f"[ ERROR ]目标文件中缺少 <{sheet_name}> 页，请检查","red")
                continue
    else:
        log_message(f"[  INFO ]{os.path.basename(source_file_path)} 中contract_name 为空或为 '/'，跳过处理","blue")
        

    #保存文件
    try:
        workbook.save(target_file_path)
    except Exception as e:
        log_message(f"[ ERROR ]无法保存目标文件: {e}","red")

    return cust_list

def update_customer_list(cust_list, target_file_path):
    # 去重后的客户列表，并排除掉 '/' 和 None 的情况
    unique_cust_list = [cust for cust in list(set(cust_list)) if cust and cust != '/']

    try:
        # 打开目标文件
        workbook = openpyxl.load_workbook(target_file_path)
    except Exception as e:
        log_message(f"[ERROR] 无法加载目标文件: {e}","red")
        return
    
    # 获取客户清单 sheet
    sheet_name = "客户清单"
    if sheet_name not in workbook.sheetnames:
        log_message(f"[ERROR] 目标文件中缺少 <{sheet_name}> 页，请检查","red")
        return

    sheet = workbook[sheet_name]

    # 清除B列从第2行开始的内容
    max_row = sheet.max_row
    for row in range(2, max_row + 1):
        sheet.cell(row=row, column=2).value = None

    # 填充去重后的客户列表从第B2行开始
    for idx, customer in enumerate(unique_cust_list, start=2):
        #print(f"正在写入第{idx}行: {customer}")  # 调试信息
        sheet.cell(row=idx, column=2, value=customer)
        # 在D列和E列填充公式
        formula_d = f"=VLOOKUP(C{idx},'附1-客户类型总览'!A:B,2,0)"
        formula_e = f"=VLOOKUP(C{idx},'附1-客户类型总览'!A:C,3,0)"
        sheet.cell(row=idx, column=4, value=formula_d)
        sheet.cell(row=idx, column=5, value=formula_e)
    
    log_message(f"[SUCCESS] 客户清单已成功写入到 <{sheet_name}> 页","green")

    # 保存文件并关闭
    try:
        workbook.save(target_file_path)
        workbook.close()
    except Exception as e:
        log_message(f"[ERROR] 无法保存目标文件: {e}","red")

#清除目标文件信息
def clear_target_file(target_file_path):
    sheet_names_to_check = ["案例清单", "通用补充信息", "数据应用类补充信息", "客户清单"]
    error_log = f"error_log.txt"
    try:
        workbook = openpyxl.load_workbook(target_file_path)
    except Exception as e:
        log_message( f"[ ERROR ]无法加载目标文件: {e}","red")
        return
    
    for sheet_name in sheet_names_to_check:
        if sheet_name in workbook.sheetnames:
            if sheet_name != "客户清单":  
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(min_row=4):
                    for cell in row:
                        cell.value = None

        else:
            log_message( f"[ ERROR ]目标文件中缺少 <{sheet_name}> 页，请检查","red")
    
    try:
        workbook.save(target_file_path)
    except Exception as e:
        log_message( f"[ ERROR ]无法保存目标文件: {e}","red")
#日志记录
def log_error(log_file,message):
    with open(log_file, 'w', encoding='gbk',errors='ignore') as f:
        f.write(message + '\n')
    print(message)

# 定义一个队列用于存储日志消息
log_queue = queue.Queue()

def log_message(message, highlight=None):
    log_queue.put((message, highlight))

def process_files_in_folder(folder_path, target_file_path):
    # 获取指定文件夹下的所有 .xlsx, .xls, .xlsm 文件
    file_patterns = ["*.xlsx", "*.xls", "*.xlsm"]
    files = []
    
    for pattern in file_patterns:
        # 获取所有匹配的文件
        matched_files = glob.glob(os.path.join(folder_path, pattern))
        # 排除以~$开头的临时文件
        matched_files = [file for file in matched_files if not os.path.basename(file).startswith('~$')]
        files.extend(matched_files)

    if not files:
        log_message("[ ERROR ] 未找到符合条件的文件.","red")
        return

    # 清理目标文件
    clear_target_file(target_file_path)
    cust_list = []
    
    # 处理每个源文件
    for source_file_path in files:  
        source_file_name = os.path.basename(source_file_path)
        log_message(f"[  INFO ] 处理文件: <{os.path.basename(source_file_name)}>","green")
        #info_display.insert(tk.END, f"[  INFO ] 处理文件: <{os.path.basename(source_file_name)}>\n","highlight")
        final_data = check_and_process_file(source_file_path)
        cust_name = load_data_to_target_file(final_data, target_file_path, source_file_path)
        cust_list.append(cust_name)

    #print(cust_list)
    # 更新客户清单
    update_customer_list(cust_list, target_file_path)

def log_message(message, highlight=None):
    log_queue.put((message, highlight))


def get_resource_path(relative_path):
    """获取打包后资源文件的路径"""
    try:
        # 获取应用程序目录
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# def load_and_display_image():
#     # 打开图片
#     image_path = "res/sunline_logo_original.png"
#     image = Image.open(image_path)
    
#     # 缩放到原来的30%
#     image = image.resize((int(image.width * 0.3), int(image.height * 0.3)), Image.ANTIALIAS)
    
#     # 转换为适合Tkinter显示的格式
#     photo = ImageTk.PhotoImage(image)
    
#     # 创建Label并显示图片
#     label = Label(root, image=photo)
#     label.image = photo  # 需要保持对图片的引用，否则图片会被垃圾回收
#     label.pack()

class App():
    def __init__(self, root):
        self.root = root
        self.root.title("案例合并工具")

        # 加载logo图片
        logo_path = get_resource_path('res/sunline_logo_original.png')
        logo_image = Image.open(logo_path)
        logo_image = logo_image.resize((int(logo_image.width * 0.4), int(logo_image.height * 0.4)))
        self.logo_photo = ImageTk.PhotoImage(logo_image)

        # 显示logo在左上角
        self.logo_label = tk.Label(self.root, image=self.logo_photo)
        self.logo_label.place(x=10, y=15)  # 设置在左上角


        self.folder_var = tk.StringVar()
        self.target_file_var = tk.StringVar()

        frame = tk.Frame(self.root)
        frame.pack(pady=30, anchor="n")

        tk.Button(frame, text="选择要合并的文件夹", command=self.select_folder, width=20).grid(row=0, column=0, padx=5, pady=5)
        tk.Button(frame, text="选择目标文件", command=self.select_target_file, width=20).grid(row=0, column=1, padx=5, pady=5)
        tk.Button(frame, text="合并案例", command=self.run_script, width=20).grid(row=1, column=0, padx=5, pady=5)
        tk.Button(frame, text="清除信息", command=self.clear_info, width=20).grid(row=1, column=1, padx=5, pady=5)

        self.info_display = scrolledtext.ScrolledText(self.root, height=15, width=80)
        self.info_display.pack(pady=5)

        self.info_label = tk.Label(self.root, text="请选择文件夹或文件", fg="red")
        self.info_label.pack()
        # 启动实时日志显示功能
        self.update_log_display()

    def update_log_display(self):
        """实时更新消息栏"""
        try:
            while True:
                message, highlight = log_queue.get_nowait()
                self.info_display.insert(tk.END, message + "\n", highlight)
                if highlight:
                    self.info_display.tag_config(highlight, foreground="white", background=highlight)
                self.info_display.see(tk.END)  # 自动滚动到最新行
        except queue.Empty:
            pass
        self.root.after(100, self.update_log_display)  # 每100ms检查一次队列

    # def log_message(self, message, color="blue"):
    #     self.info_display.insert(tk.END, f"{message}\n", "highlight")
    #     self.info_display.tag_config("highlight", foreground="white", background=color)

    def select_folder(self):
        folder_path = filedialog.askdirectory()
        self.clear_info()
        if folder_path:
            self.folder_var.set(folder_path)
            self.update_info_label(folder_path)
            log_message("[SUCCESS] 源文件夹获取成功.", "green")
        else:
            self.info_label.config(text="请选择源文件所在的文件夹", fg='red')

    def select_target_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if file_path:
            self.target_file_var.set(file_path)
            self.update_info_label(file_path)
            log_message("[SUCCESS] 目标文件获取成功.", "green")
        else:
            self.info_label.config(text="请选择目标文件", fg='red')

    def update_info_label(self, path):
        self.info_label.config(text=f"已选择：{os.path.basename(path)}", fg='green')

    def run_script(self):
        self.info_display.delete(1.0, tk.END)
        folder_path = self.folder_var.get()
        target_file_path = self.target_file_var.get()

        if not folder_path or not Path(folder_path).is_dir():
            log_message(f"[ ERROR] 请选择有效文件夹.","red")
            return
        
        if not target_file_path or not Path(target_file_path).is_file():
            log_message(f"[ ERROR] 请选择有效文件.","red")
            return

        # try:
        #     log_message(self.info_display,f"[  INFO ] |源文件路径：{folder_path}.  \n |目标文件：{target_file_path}.", "blue")
        #     rename_file(folder_path)
        #     log_message(self.info_display,"[  INFO ] 文件名处理完成.", "green")
        #     log_message(self.info_display,"[  INFO ] 正在合并请稍后.", "green")
        #     process_files_in_folder(folder_path, target_file_path)
        #     log_message(self.info_display,"[  INFO ] 处理客户清单.", "blue")
        #     log_message(self.info_display,"[  INFO ] 文件合并处理完成.", "blue")
                # 在后台线程中运行脚本
        threading.Thread(target=self.run_background_script, args=(folder_path, target_file_path)).start()

        # except Exception as e:
        #     log_message(self.info_display, f"[ ERROR] 执行脚本失败: {e}","red")
    def run_background_script(self, folder_path, target_file_path):
        try:
            log_message(f"--++本脚本最终解释权归长亮科技所有++--", "blue")
            log_message(f"[  INFO ] 源文件路径：{folder_path}\n目标文件：{target_file_path}.", "blue")
            rename_file(folder_path)
            log_message(f"[  INFO ] 文件名处理完成.", "green")
            log_message(f"[  INFO ] 正在合并请稍后.", "green")
            process_files_in_folder(folder_path, target_file_path)
            log_message(f"[  INFO ] 处理客户清单.", "green")
            log_message(f"[  INFO ] 文件合并处理完成.", "green")
        except Exception as e:
            log_message(f"[ ERROR] 执行脚本失败: {e}","red")

    def clear_info(self):
        self.folder_var.set("")
        self.target_file_var.set("")
        self.info_display.delete(1.0, tk.END)



if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()



'''
if __name__ == "__main__":
    # 设定文件夹路径和目标文件路径
    #folder_path = 'D:\github\99-0thers\case_collector\第二季度案例收集\第二季度的案例收集_bak'
    #target_file_path = 'D:/github/99-0thers/case_collector/1111.xlsx'
    #传入folder_path，target_file_path参数 并校验个数
    if len(sys.argv) != 3:
        print("Usage: python test.py <folder_path> <target_file_path>")
        sys.exit(1)

    folder_path = sys.argv[1]
    target_file_path = sys.argv[2]
    if not os.path.isdir(folder_path):
        print(f"[ ERROR ] <{folder_path}> 不是一个有效的文件夹路径。")
        sys.exit(1)

    #处理文件夹下的文件名，去除空格
    rename_file(folder_path)
    print('[  INFO ] 文件名处理完成')
    # 调用函数处理文件夹中的文件
    process_files_in_folder(folder_path, target_file_path)
'''
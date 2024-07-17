import pandas as pd
import numpy as np
import string
import os
from datetime import date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def column_index_to_string(index):
    """Converts numeric index to alphabetic column name (like Excel)."""
    string_index = ''
    while index >= 0:
        remainder = index % 26
        string_index = string.ascii_uppercase[remainder] + string_index
        index = index // 26 - 1
    return string_index

def compare_sheets(file_a, file_b, sheet_name):
    # 读取两个 Excel 文件的指定 sheet 页
    df_a = pd.read_excel(file_a, sheet_name=sheet_name)
    df_b = pd.read_excel(file_b, sheet_name=sheet_name)

    # 确保字段名的一致性
    df_a.columns = df_a.columns.str.strip()
    df_b.columns = df_b.columns.str.strip()

    # 添加辅助列来保存原始行号
    df_a['_original_index'] = df_a.index
    df_b['_original_index'] = df_b.index

    # 找出新增行和删除行的原始索引
    added_rows = df_b[~df_b['_original_index'].isin(df_a['_original_index'])]
    deleted_rows = df_a[~df_a['_original_index'].isin(df_b['_original_index'])]

    # 重置索引以便于后续操作
    df_a.reset_index(drop=True, inplace=True)
    df_b.reset_index(drop=True, inplace=True)

    # 使用一个标记列来识别差异
    df_b['_source'] = 'B'

    # 合并数据并标记差异
    merged_df = pd.concat([df_a, df_b])
    duplicates = merged_df.duplicated(subset=df_a.columns[:-2], keep=False)
    diff_df = merged_df[~duplicates]

    # 找出只在 B 中的记录
    diff_in_b = diff_df[diff_df['_source'] == 'B'].drop(columns=['_source'])

    # 获取差异项的索引
    diff_idx_b = diff_in_b.index.tolist()

    # 获取列名
    columns_b = diff_in_b.columns.tolist()

    # 定义排除的最大列索引（Y列）
    max_col_index = string.ascii_uppercase.index('Y')

    # 比较 B 表中的内容，只关注新增和删除行之外的差异，并且排除大于Y列的变更
    diff_list_b = []
    for idx in diff_idx_b:
        original_index = df_b.loc[idx, '_original_index']
        if original_index not in deleted_rows['_original_index'].values:
            for col in columns_b:
                col_index = df_b.columns.get_loc(col)
                if col_index > max_col_index:
                    continue
                old_value = df_a.loc[idx, col] if original_index in df_a['_original_index'].values else np.nan
                new_value = df_b.loc[idx, col]
                if pd.isna(old_value) and pd.isna(new_value):
                    continue  # 忽略两边都是 NaN 的情况
                if old_value != new_value:
                    col_str = column_index_to_string(col_index)
                    diff_list_b.append((original_index + 1, col_str, old_value, new_value))

    # 新增行
    new_rows = [(f"新增行: 第{idx + 1}行: {row.drop(['_original_index']).tolist()}") for idx, row in added_rows.iterrows()]

    # 删除行
    removed_rows = [(f"删除行: 第{idx + 1}行: {row.drop(['_original_index']).tolist()}") for idx, row in deleted_rows.iterrows()]

    # 读取 file_a 和 file_b 中的信息
    table_english_name_b = df_b.at[5, df_b.columns[2]]
    table_chinese_name_b = df_b.at[6, df_b.columns[2]]
    table_english_name_a = df_a.at[5, df_a.columns[2]]
    table_chinese_name_a = df_a.at[6, df_a.columns[2]]

    # 检查表英文名和表中文名是否变更
    name_changes = []
    if table_english_name_b != table_english_name_a:
        name_changes.append(f"英文表名变更: {table_english_name_a} -> {table_english_name_b}")
    if table_chinese_name_b != table_chinese_name_a:
        name_changes.append(f"中文表名变更: {table_chinese_name_a} -> {table_chinese_name_b}")

    diff_result = {
        "Only in B": [(f"第{idx}行第{col}: {old_value}->{new_value}") for (idx, col, old_value, new_value) in diff_list_b],
        "New Rows": new_rows,
        "Removed Rows": removed_rows,
        "Name Changes": name_changes if name_changes else ["无变更"]
    }

    return diff_result

def write_comparison_to_excel(file_a, file_b, sheet_name, diff_result):
    output_file = "聚合层设计文档变更记录对比.xlsx"
    today_str = date.today().strftime('%Y-%m-%d')
    
    # 读取 file_b 中的信息
    df_b_info = pd.read_excel(file_b, sheet_name=sheet_name)
    table_english_name = df_b_info.at[5, df_b_info.columns[2]]
    table_chinese_name = df_b_info.at[6, df_b_info.columns[2]]

    # 检查文件是否存在
    if os.path.exists(output_file):
        # 如果文件存在，打开文件
        book = load_workbook(output_file)
        if "Sheet1" in book.sheetnames:
            sheet = book["Sheet1"]
        else:
            sheet = book.create_sheet("Sheet1")
    else:
        # 如果文件不存在，新建文件
        book = Workbook()
        sheet = book.active
        sheet.title = "Sheet1"

        # 写入表头
        sheet.append(["文件名", "表英文名", "表中文名", "变更记录", "新增行", "删除行", "对比日期", "表名变更"])

    # 删除已有的相同数据
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == file_b and row[6].value == today_str:
            sheet.delete_rows(row[0].row, 1)

    # 写入新数据
    changes_only_in_b = "\n".join(diff_result["Only in B"]) if diff_result["Only in B"] else "无变更"
    new_rows = "\n".join(diff_result["New Rows"]) if diff_result["New Rows"] else "无变更"
    removed_rows = "\n".join(diff_result["Removed Rows"]) if diff_result["Removed Rows"] else "无变更"
    name_changes = "\n".join(diff_result["Name Changes"])

    sheet.append([file_b, table_english_name, table_chinese_name, changes_only_in_b, new_rows, removed_rows, today_str, name_changes])

    # 自动调整列宽和设置对齐方式
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) if (max_length + 2) < 25 else 25
        sheet.column_dimensions[col_letter].width = adjusted_width
        for cell in col:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # 保存文件
    book.save(output_file)

# 示例用法
if __name__ == "__main__":
    file_a = 'path/to/your/old_file_A.xlsx'
    file_b = 'path/to/your/new_file_B.xlsx'
    sheet_name = '6.表设计'  # 指定要比较的 sheet 页名称

    diff_result = compare_sheets(file_a, file_b, sheet_name)
    write_comparison_to_excel(file_a, file_b, sheet_name, diff_result)

    print("Comparison completed and results written to Excel.")

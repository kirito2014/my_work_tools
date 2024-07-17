import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import date
import os

# 提取并处理数据框
def process_dataframe(df):
    df = df.iloc[4:]  # 从第四行开始
    df.columns = ['group_no', 'ser_no'] + [f'columns_{i+1}' for i in range(len(df.columns) - 2)]
    df.reset_index(drop=True, inplace=True)
    return df

# 找到新增和删除的字段
def find_added_and_deleted_columns(df_a, df_b):
    # 合并数据框，找到新增和删除的字段
    merged_a = df_b.merge(df_a, on='columns_2', how='left', indicator=True)
    added_rows = merged_a[merged_a['_merge'] == 'left_only'].drop(columns=['_merge'])
    added_rows.reset_index(inplace=True)
    added_rows = added_rows[['columns_2', 'columns_1', 'group_no', 'ser_no']]

    merged_b = df_a.merge(df_b, on='columns_2', how='left', indicator=True)
    deleted_rows = merged_b[merged_b['_merge'] == 'left_only'].drop(columns=['_merge'])
    deleted_rows.reset_index(inplace=True)
    deleted_rows = deleted_rows[['columns_2', 'columns_1', 'group_no', 'ser_no']]

    return added_rows, deleted_rows

# 比较两个数据框的字段并记录变更信息
def compare_sheets(file_a, file_b, sheet_name):
    # 读取 Excel 文件
    df_a = pd.read_excel(file_a, sheet_name=sheet_name)
    df_b = pd.read_excel(file_b, sheet_name=sheet_name)

    # 处理数据框
    df_a_processed = process_dataframe(df_a)
    df_b_processed = process_dataframe(df_b)

    # 找到新增和删除的字段
    added_rows, deleted_rows = find_added_and_deleted_columns(df_a_processed, df_b_processed)

    # 从 df_a 和 df_b 中去除对应的新增和删除行
    df_a_filtered = df_a_processed[~df_a_processed['columns_2'].isin(deleted_rows['columns_2'])]
    df_b_filtered = df_b_processed[~df_b_processed['columns_2'].isin(added_rows['columns_2'])]

    # 合并过滤后的数据框，并去除重复数据
    df_a_filtered['_source'] = 'A'
    df_b_filtered['_source'] = 'B'
    merged = pd.concat([df_a_filtered, df_b_filtered])
    merged = merged.drop_duplicates(subset=['columns_2', 'columns_1', 'columns_3', 'columns_n'], keep=False)

    # 获取变更的字段详细信息
    diff_df = merged.loc[(merged.filter(like='_A') != merged.filter(like='_B')).any(axis=1)]
    diff_list_b = []
    for idx, row in diff_df.iterrows():
        for col in df_a_filtered.columns[2:]:
            if row[f'{col}_A'] != row[f'{col}_B']:
                diff_list_b.append((row['columns_2'], col, row[f'{col}_A'], row[f'{col}_B']))

    new_fields = [f"新增字段: {row['columns_2']} (group_no: {row['group_no']}, ser_no: {row['ser_no']})" for _, row in added_rows.iterrows()]
    removed_fields = [f"删除字段: {row['columns_2']} (group_no: {row['group_no']}, ser_no: {row['ser_no']})" for _, row in deleted_rows.iterrows()]

    diff_result = {
        "Only in B": [f"第{col}列第{df_a_filtered.columns.get_loc(col) + 1}列: {old_value}->{new_value}" for (col, old_value, new_value) in diff_list_b],
        "New Fields": new_fields,
        "Removed Fields": removed_fields
    }

    return diff_result

def write_comparison_to_excel(file_a, file_b, sheet_name, diff_result):
    output_file = "聚合层设计文档变更记录对比.xlsx"
    today_str = date.today().strftime('%Y-%m-%d')
    
    # 读取 file_b 中的信息
    df_b_info = pd.read_excel(file_b, sheet_name=sheet_name)
    table_english_name = df_b_info.at[5, df_b_info.columns[2]]
    table_chinese_name = df_b_info.at[6, df_b_info.columns[2]]

    # 检查文件是否存在
    if os.path.exists(output_file):
        # 如果文件存在，打开文件
        book = load_workbook(output_file)
        if "Sheet1" in book.sheetnames:
            sheet = book["Sheet1"]
        else:
            sheet = book.create_sheet("Sheet1")
    else:
        # 如果文件不存在，新建文件
        book = Workbook()
        sheet = book.active
        sheet.title = "Sheet1"

        # 写入表头
        sheet.append(["文件名", "表英文名", "表中文名", "变更记录", "新增字段", "删除字段", "对比日期"])

    # 删除已有的相同数据
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == file_b and row[6].value == today_str:
            sheet.delete_rows(row[0].row, 1)

    # 写入新数据
    changes_only_in_b = "\n".join(diff_result["Only in B"]) if diff_result["Only in B"] else "无变更"
    new_fields = "\n".join(diff_result["New Fields"]) if diff_result["New Fields"] else "无变更"
    removed_fields = "\n".join(diff_result["Removed Fields"]) if diff_result["Removed Fields"] else "无变更"

    sheet.append([file_b, table_english_name, table_chinese_name, changes_only_in_b, new_fields, removed_fields, today_str])

    # 自动调整列宽和设置对齐方式
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[col_letter].width = adjusted_width
        for cell in col:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # 保存文件
    book.save(output_file)

# 示例用法
if __name__ == "__main__":
    file_a = 'path/to/your/old_file_A.xlsx'
    file_b = 'path/to/your/new_file_B.xlsx'
    sheet_name = '6.表设计'  # 指定要比较的 sheet 页名称

    diff_result = compare_sheets(file_a, file_b, sheet_name)
    write_comparison_to_excel(file_a, file_b, sheet_name, diff_result)

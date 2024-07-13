import pandas as pd
from openpyxl import load_workbook

def copy_sheets_and_metadata(source_file, target_file):
    # 加载源Excel文件
    src_wb = load_workbook(filename=source_file, data_only=True)
    src_index_sheet = src_wb['index']
    
    # 加载目标Excel文件
    tgt_wb = load_workbook(filename=target_file)
    
    # 初始化错误列表
    error_list = []
    
    # 获取Index表中所有标记为Y的行
    index_data = pd.read_excel(source_file, sheet_name='index', usecols="C,D,M")
    index_data = index_data.iloc[1:]
    index_data.columns = ['table_id','table_name','enable_flag']
    index_data['table_id'] = index_data['table_id'].apply(lambda x: x.split("'")[1] if "'" in x else x)
    filtered_index = index_data[index_data['enable_flag'] == 'Y']
    
    for _, row in filtered_index.iterrows():
        table_name = row['table_name']
        table_id = row['table_id']
        
        # 检查源文件中是否存在对应的sheet页
        if sheet_name not in src_wb.sheetnames:
            error_list.append(f"Sheet {sheet_name} not found in source file.")
            continue
        
        # 处理目标文件中的sheet页
        if sheet_name in tgt_wb.sheetnames:
            del tgt_wb[sheet_name]
        tgt_wb.create_sheet(sheet_name)
    
    for _, row in filtered_index.iterrows():
        sheet_name = row['Sheet Name']
        table_id = row['Table ID']
        
        # 检查源文件中是否存在对应的sheet页
        if sheet_name not in src_wb.sheetnames:
            error_list.append(f"Sheet {sheet_name} not found in source file.")
            continue
        
        # 处理目标文件中的sheet页
        if sheet_name in tgt_wb.sheetnames:
            del tgt_wb[sheet_name]
        tgt_wb.create_sheet(sheet_name)
        
        # 复制sheet页内容
        src_sheet = src_wb[sheet_name]
        tgt_sheet = tgt_wb[sheet_name]
        
        for row in src_sheet.iter_rows(values_only=True):
            tgt_sheet.append(row)
        
        # 复制数据字典
        rem_data_dict_sheet = 'rem-数据字典'
        if rem_data_dict_sheet not in src_wb.sheetnames:
            error_list.append(f"Data dictionary sheet not found for table {table_id}.")
        else:
            src_dd_sheet = src_wb[rem_data_dict_sheet]
            tgt_dd_sheet = tgt_wb[rem_data_dict_sheet] if rem_data_dict_sheet in tgt_wb.sheetnames else tgt_wb.create_sheet(rem_data_dict_sheet)
            
            data_dict_found = False
            for row in src_dd_sheet.iter_rows(min_row=2, values_only=True):
                if row[1] == table_id:
                    data_dict_found = True
                    tgt_dd_sheet.append(row)
            
            if not data_dict_found:
                error_list.append(f"Data dictionary not found for table {table_id}.")
        
        # 复制代码映射
        rem_code_map_sheet = 'rem-代码映射'
        if rem_code_map_sheet not in src_wb.sheetnames:
            error_list.append(f"Code mapping sheet not found for table {table_id}.")
        else:
            src_cm_sheet = src_wb[rem_code_map_sheet]
            tgt_cm_sheet = tgt_wb[rem_code_map_sheet] if rem_code_map_sheet in tgt_wb.sheetnames else tgt_wb.create_sheet(rem_code_map_sheet)
            
            code_map_found = False
            for row in src_cm_sheet.iter_rows(min_row=3, values_only=True):
                if row[8] == table_id:
                    code_map_found = True
                    tgt_cm_sheet.append(row)
            
            if not code_map_found:
                error_list.append(f"Code mapping not found for table {table_id}.")
    
    # 保存目标文件
    tgt_wb.save(target_file)
    
    # 返回错误列表
    return error_list

# 使用示例
source_file = 'source.xlsx'
target_file = 'target.xlsx'
errors = copy_sheets_and_metadata(source_file, target_file)
print(f"Errors: {errors}")

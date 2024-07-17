import os
import re
import openpyxl
import sys

# 获取当前文件所在目录作为 ETL_HOME
ETL_HOME = os.path.dirname(os.path.abspath(__file__))

# 函数用于根据文件名判断 belong_theme 的值
def get_belong_theme(file_name):
    # 去掉文件名的扩展名并转换为大写形式
    file_name_stripped = os.path.splitext(file_name)[0].upper()
    # 根据特定规则判断 belong_theme 的值
    if file_name_stripped[3:6] == "S01":
        return "AGL_CORP"
    elif file_name_stripped[3:6] == "S02":
        return "AGL_RTLB"
    elif file_name_stripped[3:6] == "S03":
        return "AGL_LOAN"
    elif file_name_stripped[3:6] == "S04":
        return "AGL_ASSM"
    elif file_name_stripped[3:6] == "S05":
        return "AGL_FINN"
    elif file_name_stripped[3:6] == "S06":
        return "AGL_OPRS"
    elif file_name_stripped[3:6] in ["S07", "S08", "S09", "S10", "S11", "S12"]:
        return "AGL_COMM"
    else:
        return "Unknown"  # 若无匹配规则，则返回 Unknown

# 函数用于提取 SQL 文件中的表名
def extract_table_names(sql_content):
    # 使用正则表达式匹配 FROM 和 JOIN 后的表名
    pattern = r'(?:FROM|JOIN)\s+(\w+\.\w+)\s+'
    matches = re.findall(pattern, sql_content)
    return matches

# 函数用于过滤表名列表中包含文件名的表名，并且排除以 "AGL_" 开头的表名
def filter_table_names(file_name, table_names):
    filtered_table_names = []
    for table_name in table_names:
        # 将表名转换为大写形式并去除空格
        table_name_cleaned = table_name.upper().replace(' ', '')
        # 排除以 "AGL_" 开头的表名
        if not table_name_cleaned.startswith("AGL_"):
            # 将处理后的表名添加到列表中
            filtered_table_names.append(table_name_cleaned)
    # 对过滤后的表名列表进行去重
    filtered_table_names = list(set(filtered_table_names))
    return filtered_table_names

# 遍历文件夹下的每个 SQL 文件
def process_folder(folder_path):
    data = []  # 存储文件名、belong_theme、表名、source_schema 和 source_table 的列表
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".sql"):
            file_path = os.path.join(folder_path, file_name)
            with open(file_path, 'r', encoding='gbk') as file:
                sql_content = file.read()
                table_names = extract_table_names(sql_content)
                filtered_table_names = filter_table_names(file_name, table_names)
                belong_theme = get_belong_theme(file_name)
                for table_name in filtered_table_names:
                    source_schema, source_table = table_name.split('.')  # 分割表名
                    source_table += '_PC'  # 拼接表名
                    data.append((file_name, belong_theme, table_name, source_schema, source_table))
    return data

# 写入数据到 Excel 文件
def write_to_excel(data, output_file):
    output_path = os.path.join(ETL_HOME, output_file)  # 拼接输出文件路径
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = '文件名'
    sheet['B1'] = 'belong_theme'
    sheet['C1'] = '表名'
    sheet['D1'] = 'source_schema'
    sheet['E1'] = 'source_table'

    # 写入数据
    for idx, (file_name, belong_theme, table_name, source_schema, source_table) in enumerate(data, start=2):
        sheet.cell(row=idx, column=1, value=file_name)
        sheet.cell(row=idx, column=2, value=belong_theme)
        sheet.cell(row=idx, column=3, value=table_name)
        sheet.cell(row=idx, column=4, value=source_schema)
        sheet.cell(row=idx, column=5, value=source_table)

    # 调整列宽
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 30

    workbook.save(output_path)

# 主函数
def main(folder_path, output_file):
    data = process_folder(folder_path)
    write_to_excel(data, output_file)

# 示例用法
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python gen.py <folder_path> <output_file>")
        sys.exit(1)
    folder_path = sys.argv[1]  # 获取文件夹路径参数
    output_file = sys.argv[2]  # 获取输出文件名参数
    main(folder_path, output_file)
